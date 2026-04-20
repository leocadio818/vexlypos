"""
Reportes estilo Revel — formato XLSX con 2 hojas (RESUMEN + DETALLE).
Tipos soportados:
  - 607 (ventas DGII)
  - 606 (compras DGII)
  - 608 (anulaciones/NC)
  - ventas-generales (todas las ventas del período con detalles completos)

Todas las hojas incluyen columna ESTADO con 6 posibles valores:
Aprobada, Rechazada, Contingencia, Cont. Manual, Procesando, Pendiente.
"""
from datetime import datetime, timezone, timedelta
from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import os
from motor.motor_asyncio import AsyncIOMotorClient
from routers.auth import get_current_user

_client = AsyncIOMotorClient(os.environ['MONGO_URL'])
db = _client[os.environ.get('DB_NAME', 'pos_db')]

router = APIRouter(prefix="/reports/revel", tags=["reports-revel"])


# ═══════════════════════════════════════════════════════════════
# Shared constants
# ═══════════════════════════════════════════════════════════════

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=16, color="1F4E78")
SUBTITLE_FONT = Font(bold=True, size=11, color="666666")
TOTAL_FILL = PatternFill("solid", fgColor="FFE699")
TOTAL_FONT = Font(bold=True, size=11)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Tier colors for ESTADO column
ESTADO_COLORS = {
    "Aprobada": "C6EFCE",        # light green
    "Rechazada": "FFC7CE",       # light red
    "Contingencia": "FFEB9C",    # light amber
    "Cont. Manual": "D9D9D9",    # gray
    "Procesando": "BDD7EE",      # light blue
    "Pendiente": "E7E6E6",       # very light gray
}
ESTADO_TEXT_COLORS = {
    "Aprobada": "006100",
    "Rechazada": "9C0006",
    "Contingencia": "9C5700",
    "Cont. Manual": "404040",
    "Procesando": "1F4E78",
    "Pendiente": "595959",
}


def _classify_estado(bill: dict) -> str:
    """Map bill ecf_status → Spanish human-readable estado."""
    status = (bill.get("ecf_status") or "").upper()
    # Manual contingency flags
    ecf_err = (bill.get("ecf_error") or "").lower()
    is_manual = (
        "contingencia_manual" in ecf_err
        or bill.get("force_contingency") is True
        or any(
            (p or {}).get("skip_ecf") is True or (p or {}).get("force_contingency") is True
            for p in (bill.get("payments") or [])
        )
    )
    if status == "FINISHED":
        return "Aprobada"
    if status == "REJECTED" or bill.get("ecf_reject_reason"):
        return "Rechazada"
    if status == "CONTINGENCIA":
        return "Cont. Manual" if is_manual else "Contingencia"
    if status == "PROCESSING":
        return "Procesando"
    if status == "REGISTERED":
        return "Procesando"
    return "Pendiente"


def _payment_amount(bill: dict, method_keywords: list) -> float:
    """Sum payment amounts that match any of the method keywords (case-insensitive)."""
    total = 0.0
    for p in (bill.get("payments") or []):
        m = str((p or {}).get("method") or "").lower()
        if any(k in m for k in method_keywords):
            total += float((p or {}).get("amount") or 0)
    return round(total, 2)


def _build_range_query(date_from: str, date_to: str) -> dict:
    """Build paid_at range query using America/Santo_Domingo timezone."""
    from zoneinfo import ZoneInfo
    from utils.timezone import get_system_timezone_name
    # This must be awaited by caller if async; here we use static fallback for simplicity.
    tz = ZoneInfo("America/Santo_Domingo")
    try:
        start = datetime.strptime(date_from, "%Y-%m-%d").replace(tzinfo=tz)
        end = datetime.strptime(date_to, "%Y-%m-%d").replace(hour=23, minute=59, second=59, tzinfo=tz)
        return {"$gte": start.astimezone(timezone.utc).isoformat(),
                "$lte": end.astimezone(timezone.utc).isoformat()}
    except Exception:
        return {"$gte": date_from, "$lte": f"{date_to}T23:59:59"}


def _apply_header(ws, row: int, headers: list):
    for col_idx, title in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def _autosize(ws, min_width: int = 10, max_width: int = 35):
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = min_width
        for c in col_cells:
            v = c.value
            if v is None:
                continue
            try:
                ln = len(str(v))
            except Exception:
                ln = 10
            if ln > max_len:
                max_len = ln
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


def _write_title_block(ws, title: str, subtitle: str, period_label: str, business_name: str):
    """Write a 4-row header: title, business, period, blank."""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    ws.cell(row=1, column=1, value=title).font = TITLE_FONT
    ws.cell(row=1, column=1).alignment = CENTER
    ws.row_dimensions[1].height = 28

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=10)
    ws.cell(row=2, column=1, value=business_name).font = SUBTITLE_FONT
    ws.cell(row=2, column=1).alignment = CENTER

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=10)
    ws.cell(row=3, column=1, value=f"{subtitle} · {period_label}").font = SUBTITLE_FONT
    ws.cell(row=3, column=1).alignment = CENTER


def _paint_estado_cell(cell, estado: str):
    cell.fill = PatternFill("solid", fgColor=ESTADO_COLORS.get(estado, "FFFFFF"))
    cell.font = Font(bold=True, color=ESTADO_TEXT_COLORS.get(estado, "000000"), size=10)
    cell.alignment = CENTER


# ═══════════════════════════════════════════════════════════════
# Report builders
# ═══════════════════════════════════════════════════════════════

async def _fetch_sales_bills(date_from: str, date_to: str) -> list:
    """Fetch all paid bills in range."""
    paid_range = _build_range_query(date_from, date_to)
    bills = await db.bills.find(
        {"paid_at": paid_range},
        {"_id": 0}
    ).sort("paid_at", 1).to_list(5000)
    return bills


async def _fetch_credit_notes(date_from: str, date_to: str) -> list:
    paid_range = _build_range_query(date_from, date_to)
    notes = await db.credit_notes.find(
        {"created_at": paid_range},
        {"_id": 0}
    ).sort("created_at", 1).to_list(2000)
    return notes


async def _fetch_expenses(date_from: str, date_to: str) -> list:
    paid_range = _build_range_query(date_from, date_to)
    cursor = db.expenses.find(
        {"expense_date": paid_range},
        {"_id": 0}
    )
    try:
        return await cursor.sort("expense_date", 1).to_list(5000)
    except Exception:
        return []


async def _get_business_info() -> dict:
    cfg = await db.system_config.find_one({}, {"_id": 0}) or {}
    return {
        "name": cfg.get("ticket_business_name") or cfg.get("restaurant_name") or "VexlyPOS",
        "rnc": cfg.get("ticket_rnc") or cfg.get("rnc") or "",
    }


def _build_ventas_workbook(title: str, subtitle: str, bills: list,
                            period_label: str, business: dict,
                            include_operational: bool = True) -> BytesIO:
    """
    Build XLSX with RESUMEN + DETALLE sheets for sales bills.
    Columns: # Factura, e-NCF, Tipo Venta, Nombre Comercial, RNC, Neto, ITBIS,
             LEY 10%, Total, Efectivo, Tarjeta Visa, Tarjeta MC, AMEX,
             Transferencia, Fecha, ESTADO [+ Cajero, Mesero, Mesa, Hora]
    """
    wb = Workbook()
    # ── Sheet 1: RESUMEN
    ws = wb.active
    ws.title = "RESUMEN"

    _write_title_block(ws, title, subtitle, period_label, business["name"])
    ws.cell(row=4, column=1, value=f"RNC: {business['rnc']}").font = SUBTITLE_FONT

    # Totals computation
    total_count = len(bills)
    by_estado = {e: 0 for e in ["Aprobada", "Rechazada", "Contingencia", "Cont. Manual", "Procesando", "Pendiente"]}
    total_neto = total_itbis = total_ley = total_total = 0.0
    total_efectivo = total_visa = total_mc = total_amex = total_transfer = 0.0
    by_tipo = {}
    for b in bills:
        e = _classify_estado(b)
        by_estado[e] = by_estado.get(e, 0) + 1
        total_neto += float(b.get("subtotal") or 0)
        total_itbis += float(b.get("itbis") or 0)
        total_ley += float(b.get("tip") or b.get("service_charge") or 0)
        total_total += float(b.get("total") or 0)
        total_efectivo += _payment_amount(b, ["efectivo", "cash"])
        total_visa += _payment_amount(b, ["visa"])
        total_mc += _payment_amount(b, ["master", "mc "])
        total_amex += _payment_amount(b, ["amex", "american"])
        total_transfer += _payment_amount(b, ["transfer", "transferencia"])
        tv = b.get("ecf_type") or "OTRO"
        by_tipo[tv] = by_tipo.get(tv, 0) + 1

    row = 6
    # Summary block 1: counts
    ws.cell(row=row, column=1, value="📊 RESUMEN DE FACTURAS").font = Font(bold=True, size=12, color="1F4E78")
    row += 1
    rows_data = [
        ("Total facturas emitidas", total_count, None),
        ("✓ Aprobadas por DGII", by_estado["Aprobada"], "Aprobada"),
        ("✗ Rechazadas por DGII", by_estado["Rechazada"], "Rechazada"),
        ("⚠ En Contingencia", by_estado["Contingencia"], "Contingencia"),
        ("📱 Cont. Manual (Plataforma externa)", by_estado["Cont. Manual"], "Cont. Manual"),
        ("⏳ Procesando", by_estado["Procesando"], "Procesando"),
        ("· Pendientes", by_estado["Pendiente"], "Pendiente"),
    ]
    for label, value, estado_key in rows_data:
        c1 = ws.cell(row=row, column=1, value=label)
        c2 = ws.cell(row=row, column=2, value=value)
        c1.alignment = LEFT
        c2.alignment = RIGHT
        c2.font = Font(bold=True, size=11)
        if estado_key:
            c1.fill = PatternFill("solid", fgColor=ESTADO_COLORS[estado_key])
            c1.font = Font(color=ESTADO_TEXT_COLORS[estado_key], size=10, bold=True)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="💰 TOTALES FISCALES (RD$)").font = Font(bold=True, size=12, color="1F4E78")
    row += 1
    for label, val in [
        ("Total NETO", total_neto),
        ("Total ITBIS (18%)", total_itbis),
        ("Total Propina/LEY (10%)", total_ley),
        ("TOTAL VENTAS", total_total),
    ]:
        ws.cell(row=row, column=1, value=label).alignment = LEFT
        c = ws.cell(row=row, column=2, value=round(val, 2))
        c.number_format = '#,##0.00'
        c.alignment = RIGHT
        c.font = Font(bold=True)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="💳 FORMAS DE PAGO (RD$)").font = Font(bold=True, size=12, color="1F4E78")
    row += 1
    for label, val in [
        ("Efectivo", total_efectivo),
        ("Tarjeta Visa", total_visa),
        ("Tarjeta MasterCard", total_mc),
        ("American Express (AMEX)", total_amex),
        ("Transferencia", total_transfer),
    ]:
        ws.cell(row=row, column=1, value=label).alignment = LEFT
        c = ws.cell(row=row, column=2, value=round(val, 2))
        c.number_format = '#,##0.00'
        c.alignment = RIGHT
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="📋 POR TIPO DE VENTA").font = Font(bold=True, size=12, color="1F4E78")
    row += 1
    for tipo, count in sorted(by_tipo.items()):
        ws.cell(row=row, column=1, value=tipo).alignment = LEFT
        ws.cell(row=row, column=2, value=count).alignment = RIGHT
        row += 1

    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 20

    # ── Sheet 2: DETALLE
    ws2 = wb.create_sheet("DETALLE")
    _write_title_block(ws2, title, subtitle, period_label, business["name"])

    headers = ["# Factura", "e-NCF", "Tipo Venta", "Nombre Comercial", "RNC",
               "Neto", "ITBIS", "LEY 10%", "Total",
               "Efectivo", "Tarjeta Visa", "Tarjeta MC", "AMEX", "Transferencia",
               "Fecha", "ESTADO"]
    if include_operational:
        headers += ["Cajero", "Mesero", "Mesa", "Hora"]

    _apply_header(ws2, 5, headers)

    row = 6
    for b in bills:
        estado = _classify_estado(b)
        paid_at = b.get("paid_at") or b.get("created_at") or ""
        fecha_str = ""
        hora_str = ""
        if paid_at:
            try:
                # paid_at is UTC ISO; convert to RD tz for display
                from zoneinfo import ZoneInfo
                dt = datetime.fromisoformat(str(paid_at).replace("Z", "+00:00"))
                local = dt.astimezone(ZoneInfo("America/Santo_Domingo"))
                fecha_str = local.strftime("%d/%m/%Y")
                hora_str = local.strftime("%H:%M:%S")
            except Exception:
                fecha_str = str(paid_at)[:10]

        cust = b.get("customer") or {}
        rnc = cust.get("rnc") or b.get("fiscal_id") or b.get("rnc") or ""
        razon = cust.get("razon_social") or b.get("razon_social") or cust.get("name") or ""

        values = [
            b.get("transaction_number") or b.get("fiscal_id") or "",
            b.get("ecf_encf") or b.get("ncf") or "",
            b.get("ecf_type") or "E32",
            razon,
            rnc,
            round(float(b.get("subtotal") or 0), 2),
            round(float(b.get("itbis") or 0), 2),
            round(float(b.get("tip") or b.get("service_charge") or 0), 2),
            round(float(b.get("total") or 0), 2),
            _payment_amount(b, ["efectivo", "cash"]),
            _payment_amount(b, ["visa"]),
            _payment_amount(b, ["master", "mc "]),
            _payment_amount(b, ["amex", "american"]),
            _payment_amount(b, ["transfer", "transferencia"]),
            fecha_str,
            estado,
        ]
        if include_operational:
            values += [
                b.get("cashier_name") or "",
                b.get("waiter_name") or "",
                b.get("table_number") or "",
                hora_str,
            ]

        for col_idx, val in enumerate(values, start=1):
            c = ws2.cell(row=row, column=col_idx, value=val)
            c.border = BORDER
            if col_idx in (1,):
                c.alignment = CENTER
            elif col_idx in (6, 7, 8, 9, 10, 11, 12, 13, 14):
                c.alignment = RIGHT
                c.number_format = '#,##0.00'
            elif col_idx == 16:
                _paint_estado_cell(c, estado)
            else:
                c.alignment = LEFT
        row += 1

    # Totals row
    if bills:
        tot_cell = ws2.cell(row=row, column=1, value="TOTAL")
        tot_cell.fill = TOTAL_FILL
        tot_cell.font = TOTAL_FONT
        tot_cell.alignment = CENTER
        for col_idx in range(2, len(headers) + 1):
            cell = ws2.cell(row=row, column=col_idx)
            cell.fill = TOTAL_FILL
            cell.font = TOTAL_FONT
            if col_idx in (6, 7, 8, 9, 10, 11, 12, 13, 14):
                col_letter = get_column_letter(col_idx)
                cell.value = f"=SUM({col_letter}6:{col_letter}{row - 1})"
                cell.number_format = '#,##0.00'
                cell.alignment = RIGHT

    ws2.freeze_panes = "A6"
    _autosize(ws2)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _build_compras_workbook(expenses: list, period_label: str, business: dict) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "RESUMEN"
    _write_title_block(ws, "REPORTE 606 — COMPRAS", "Gastos y compras a proveedores", period_label, business["name"])

    total_count = len(expenses)
    total_neto = total_itbis = total_total = 0.0
    by_supplier = {}
    for e in expenses:
        total_neto += float(e.get("subtotal") or 0)
        total_itbis += float(e.get("itbis") or 0)
        total_total += float(e.get("total") or 0)
        s = e.get("supplier_name") or "Sin proveedor"
        by_supplier[s] = by_supplier.get(s, 0) + float(e.get("total") or 0)

    row = 6
    ws.cell(row=row, column=1, value="📊 RESUMEN").font = Font(bold=True, size=12, color="1F4E78")
    row += 1
    for label, val, fmt in [
        ("Total compras registradas", total_count, "#,##0"),
        ("Total NETO (RD$)", round(total_neto, 2), "#,##0.00"),
        ("Total ITBIS pagado (RD$)", round(total_itbis, 2), "#,##0.00"),
        ("TOTAL GENERAL (RD$)", round(total_total, 2), "#,##0.00"),
    ]:
        ws.cell(row=row, column=1, value=label).alignment = LEFT
        c = ws.cell(row=row, column=2, value=val)
        c.alignment = RIGHT
        c.number_format = fmt
        c.font = Font(bold=True)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="🏢 POR PROVEEDOR").font = Font(bold=True, size=12, color="1F4E78")
    row += 1
    for name, amt in sorted(by_supplier.items(), key=lambda x: -x[1]):
        ws.cell(row=row, column=1, value=name).alignment = LEFT
        c = ws.cell(row=row, column=2, value=round(amt, 2))
        c.alignment = RIGHT
        c.number_format = '#,##0.00'
        row += 1

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 20

    # DETALLE
    ws2 = wb.create_sheet("DETALLE")
    _write_title_block(ws2, "REPORTE 606 — DETALLE", "Compras y gastos", period_label, business["name"])
    headers = ["# Gasto", "NCF Proveedor", "Tipo NCF", "Proveedor", "RNC Proveedor",
               "Neto", "ITBIS", "Total", "Forma Pago", "Categoría", "Fecha", "Referencia"]
    _apply_header(ws2, 5, headers)
    row = 6
    for e in expenses:
        fecha = e.get("expense_date") or ""
        fecha_str = str(fecha)[:10]
        values = [
            e.get("expense_number") or e.get("id", "")[:8],
            e.get("ncf") or "",
            e.get("ncf_type") or "",
            e.get("supplier_name") or "",
            e.get("supplier_rnc") or "",
            round(float(e.get("subtotal") or 0), 2),
            round(float(e.get("itbis") or 0), 2),
            round(float(e.get("total") or 0), 2),
            e.get("payment_method") or "",
            e.get("category") or "",
            fecha_str,
            e.get("reference") or "",
        ]
        for col_idx, val in enumerate(values, start=1):
            c = ws2.cell(row=row, column=col_idx, value=val)
            c.border = BORDER
            if col_idx in (6, 7, 8):
                c.alignment = RIGHT
                c.number_format = '#,##0.00'
            else:
                c.alignment = LEFT
        row += 1
    if expenses:
        tot_cell = ws2.cell(row=row, column=1, value="TOTAL")
        tot_cell.fill = TOTAL_FILL
        tot_cell.font = TOTAL_FONT
        tot_cell.alignment = CENTER
        for col_idx in range(2, len(headers) + 1):
            cell = ws2.cell(row=row, column=col_idx)
            cell.fill = TOTAL_FILL
            cell.font = TOTAL_FONT
            if col_idx in (6, 7, 8):
                col_letter = get_column_letter(col_idx)
                cell.value = f"=SUM({col_letter}6:{col_letter}{row - 1})"
                cell.number_format = '#,##0.00'
                cell.alignment = RIGHT
    ws2.freeze_panes = "A6"
    _autosize(ws2)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ═══════════════════════════════════════════════════════════════
# Endpoints
# ═══════════════════════════════════════════════════════════════

def _period_label(date_from: str, date_to: str) -> str:
    return f"Del {date_from} al {date_to}"


def _xlsx_response(buf: BytesIO, filename: str) -> StreamingResponse:
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/607")
async def report_607(
    date_from: str = Query(..., description="YYYY-MM-DD"),
    date_to: str = Query(..., description="YYYY-MM-DD"),
    user=Depends(get_current_user),
):
    """Reporte 607 — Ventas a DGII (formato Revel)."""
    bills = await _fetch_sales_bills(date_from, date_to)
    # 607 covers ALL sales (consumidor final + crédito fiscal + gubernamental)
    bills_607 = [b for b in bills if b.get("paid_at")]
    business = await _get_business_info()
    buf = _build_ventas_workbook(
        title="REPORTE 607 — VENTAS DGII",
        subtitle="Detalle de ventas fiscales",
        bills=bills_607,
        period_label=_period_label(date_from, date_to),
        business=business,
    )
    fname = f"607_{date_from}_al_{date_to}.xlsx"
    return _xlsx_response(buf, fname)


@router.get("/606")
async def report_606(
    date_from: str = Query(...),
    date_to: str = Query(...),
    user=Depends(get_current_user),
):
    """Reporte 606 — Compras DGII (formato Revel)."""
    expenses = await _fetch_expenses(date_from, date_to)
    business = await _get_business_info()
    buf = _build_compras_workbook(expenses, _period_label(date_from, date_to), business)
    fname = f"606_{date_from}_al_{date_to}.xlsx"
    return _xlsx_response(buf, fname)


@router.get("/608")
async def report_608(
    date_from: str = Query(...),
    date_to: str = Query(...),
    user=Depends(get_current_user),
):
    """Reporte 608 — Anulaciones / Notas de Crédito (formato Revel)."""
    bills = await _fetch_sales_bills(date_from, date_to)
    notes = await _fetch_credit_notes(date_from, date_to)
    business = await _get_business_info()

    # Build "virtual bills" from credit_notes + REJECTED bills
    virtual = []
    for n in notes:
        virtual.append({
            "transaction_number": n.get("transaction_number") or n.get("id", "")[:8],
            "ecf_encf": n.get("encf") or n.get("ncf") or "",
            "ecf_type": "E34",
            "customer": {"rnc": n.get("customer_rnc") or "", "razon_social": n.get("customer_name") or ""},
            "subtotal": n.get("subtotal") or 0,
            "itbis": n.get("itbis") or 0,
            "tip": 0,
            "total": n.get("total") or 0,
            "payments": [],
            "paid_at": n.get("created_at"),
            "ecf_status": n.get("ecf_status") or "FINISHED",
            "cashier_name": n.get("created_by_name") or "",
            "ecf_reject_reason": n.get("ecf_reject_reason"),
        })
    rejected = [b for b in bills if (b.get("ecf_status") or "").upper() == "REJECTED"]
    combined = virtual + rejected

    buf = _build_ventas_workbook(
        title="REPORTE 608 — ANULACIONES",
        subtitle="Notas de Crédito (E34) + Facturas Rechazadas",
        bills=combined,
        period_label=_period_label(date_from, date_to),
        business=business,
    )
    fname = f"608_{date_from}_al_{date_to}.xlsx"
    return _xlsx_response(buf, fname)


@router.get("/ventas-generales")
async def report_ventas_generales(
    date_from: str = Query(...),
    date_to: str = Query(...),
    user=Depends(get_current_user),
):
    """Reporte General de Ventas — Todas las facturas del período con detalles operacionales."""
    bills = await _fetch_sales_bills(date_from, date_to)
    business = await _get_business_info()
    buf = _build_ventas_workbook(
        title="REPORTE GENERAL DE VENTAS",
        subtitle="Todas las facturas del período",
        bills=bills,
        period_label=_period_label(date_from, date_to),
        business=business,
        include_operational=True,
    )
    fname = f"ventas_generales_{date_from}_al_{date_to}.xlsx"
    return _xlsx_response(buf, fname)
