"""
Reportes XLSX detallados — formato XLSX con 2 hojas (RESUMEN + DETALLE).
Tipos soportados:
  - 607 (ventas DGII)
  - 606 (compras DGII)
  - 608 (anulaciones/NC)
  - ventas-generales (todas las ventas del período con detalles completos)

Todas las hojas incluyen columna ESTADO con 6 posibles valores:
Aprobada, Rechazada, Contingencia, Cont. Manual, Procesando, Pendiente.
"""
from datetime import datetime, timezone, timedelta
from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import os
from motor.motor_asyncio import AsyncIOMotorClient
from routers.auth import get_current_user

_client = AsyncIOMotorClient(os.environ['MONGO_URL'])
db = _client[os.environ.get('DB_NAME', 'pos_db')]

router = APIRouter(prefix="/reports/xlsx", tags=["reports-xlsx"])


# ═══════════════════════════════════════════════════════════════
# Shared constants
# ═══════════════════════════════════════════════════════════════

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=16, color="1F4E78")
SUBTITLE_FONT = Font(bold=True, size=11, color="666666")
TOTAL_FILL = PatternFill("solid", fgColor="FFE699")
TOTAL_FONT = Font(bold=True, size=11)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Tier colors for ESTADO column
ESTADO_COLORS = {
    "Aprobada": "C6EFCE",        # light green
    "Rechazada": "FFC7CE",       # light red
    "Contingencia": "FFEB9C",    # light amber
    "Cont. Manual": "D9D9D9",    # gray
    "Procesando": "BDD7EE",      # light blue
    "Pendiente": "E7E6E6",       # very light gray
}
ESTADO_TEXT_COLORS = {
    "Aprobada": "006100",
    "Rechazada": "9C0006",
    "Contingencia": "9C5700",
    "Cont. Manual": "404040",
    "Procesando": "1F4E78",
    "Pendiente": "595959",
}


def _classify_estado(bill: dict) -> str:
    """Map bill ecf_status → Spanish human-readable estado."""
    status = (bill.get("ecf_status") or "").upper()
    # Manual contingency flags
    ecf_err = (bill.get("ecf_error") or "").lower()
    is_manual = (
        "contingencia_manual" in ecf_err
        or bill.get("force_contingency") is True
        or any(
            (p or {}).get("skip_ecf") is True or (p or {}).get("force_contingency") is True
            for p in (bill.get("payments") or [])
        )
    )
    if status == "FINISHED":
        return "Aprobada"
    if status == "REJECTED" or bill.get("ecf_reject_reason"):
        return "Rechazada"
    if status == "CONTINGENCIA":
        return "Cont. Manual" if is_manual else "Contingencia"
    if status == "PROCESSING":
        return "Procesando"
    if status == "REGISTERED":
        return "Procesando"
    return "Pendiente"


def _payment_amount(bill: dict, method_keywords: list) -> float:
    """Sum payment amounts that match any of the method keywords (case-insensitive)."""
    total = 0.0
    for p in (bill.get("payments") or []):
        m = str((p or {}).get("method") or "").lower()
        if any(k in m for k in method_keywords):
            total += float((p or {}).get("amount") or 0)
    return round(total, 2)


def _build_range_query(date_from: str, date_to: str) -> dict:
    """Build paid_at range query using America/Santo_Domingo timezone."""
    from zoneinfo import ZoneInfo
    from utils.timezone import get_system_timezone_name
    # This must be awaited by caller if async; here we use static fallback for simplicity.
    tz = ZoneInfo("America/Santo_Domingo")
    try:
        start = datetime.strptime(date_from, "%Y-%m-%d").replace(tzinfo=tz)
        end = datetime.strptime(date_to, "%Y-%m-%d").replace(hour=23, minute=59, second=59, tzinfo=tz)
        return {"$gte": start.astimezone(timezone.utc).isoformat(),
                "$lte": end.astimezone(timezone.utc).isoformat()}
    except Exception:
        return {"$gte": date_from, "$lte": f"{date_to}T23:59:59"}


def _apply_header(ws, row: int, headers: list):
    for col_idx, title in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def _autosize(ws, min_width: int = 10, max_width: int = 35):
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = min_width
        for c in col_cells:
            v = c.value
            if v is None:
                continue
            try:
                ln = len(str(v))
            except Exception:
                ln = 10
            if ln > max_len:
                max_len = ln
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


def _write_title_block(ws, title: str, subtitle: str, period_label: str, business_name: str):
    """Write a 4-row header: title, business, period, blank."""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    ws.cell(row=1, column=1, value=title).font = TITLE_FONT
    ws.cell(row=1, column=1).alignment = CENTER
    ws.row_dimensions[1].height = 28

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=10)
    ws.cell(row=2, column=1, value=business_name).font = SUBTITLE_FONT
    ws.cell(row=2, column=1).alignment = CENTER

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=10)
    ws.cell(row=3, column=1, value=f"{subtitle} · {period_label}").font = SUBTITLE_FONT
    ws.cell(row=3, column=1).alignment = CENTER


def _paint_estado_cell(cell, estado: str):
    cell.fill = PatternFill("solid", fgColor=ESTADO_COLORS.get(estado, "FFFFFF"))
    cell.font = Font(bold=True, color=ESTADO_TEXT_COLORS.get(estado, "000000"), size=10)
    cell.alignment = CENTER


# ═══════════════════════════════════════════════════════════════
# Report builders
# ═══════════════════════════════════════════════════════════════

async def _fetch_sales_bills(date_from: str, date_to: str) -> list:
    """Fetch all paid bills in range."""
    paid_range = _build_range_query(date_from, date_to)
    bills = await db.bills.find(
        {"paid_at": paid_range},
        {"_id": 0}
    ).sort("paid_at", 1).to_list(5000)
    return bills


async def _fetch_credit_notes(date_from: str, date_to: str) -> list:
    paid_range = _build_range_query(date_from, date_to)
    notes = await db.credit_notes.find(
        {"created_at": paid_range},
        {"_id": 0}
    ).sort("created_at", 1).to_list(2000)
    return notes


async def _fetch_expenses(date_from: str, date_to: str) -> list:
    paid_range = _build_range_query(date_from, date_to)
    cursor = db.expenses.find(
        {"expense_date": paid_range},
        {"_id": 0}
    )
    try:
        return await cursor.sort("expense_date", 1).to_list(5000)
    except Exception:
        return []


async def _get_business_info() -> dict:
    cfg = await db.system_config.find_one({}, {"_id": 0}) or {}
    return {
        "name": cfg.get("ticket_business_name") or cfg.get("restaurant_name") or "VexlyPOS",
        "rnc": cfg.get("ticket_rnc") or cfg.get("rnc") or "",
    }


def _build_ventas_workbook(title: str, subtitle: str, bills: list,
                            period_label: str, business: dict,
                            include_operational: bool = True) -> BytesIO:
    """
    Build XLSX with RESUMEN + DETALLE sheets for sales bills.
    Columns: # Factura, e-NCF, Tipo Venta, Nombre Comercial, RNC, Neto, ITBIS,
             LEY 10%, Total, Efectivo, Tarjeta Visa, Tarjeta MC, AMEX,
             Transferencia, Fecha, ESTADO [+ Cajero, Mesero, Mesa, Hora]
    """
    wb = Workbook()
    # ── Sheet 1: RESUMEN
    ws = wb.active
    ws.title = "RESUMEN"

    _write_title_block(ws, title, subtitle, period_label, business["name"])
    ws.cell(row=4, column=1, value=f"RNC: {business['rnc']}").font = SUBTITLE_FONT

    # Totals computation
    total_count = len(bills)
    by_estado = {e: 0 for e in ["Aprobada", "Rechazada", "Contingencia", "Cont. Manual", "Procesando", "Pendiente"]}
    total_neto = total_itbis = total_ley = total_total = 0.0
    total_efectivo = total_visa = total_mc = total_amex = total_transfer = 0.0
    by_tipo = {}
    for b in bills:
        e = _classify_estado(b)
        by_estado[e] = by_estado.get(e, 0) + 1
        total_neto += float(b.get("subtotal") or 0)
        total_itbis += float(b.get("itbis") or 0)
        total_ley += float(b.get("tip") or b.get("service_charge") or 0)
        total_total += float(b.get("total") or 0)
        total_efectivo += _payment_amount(b, ["efectivo", "cash"])
        total_visa += _payment_amount(b, ["visa"])
        total_mc += _payment_amount(b, ["master", "mc "])
        total_amex += _payment_amount(b, ["amex", "american"])
        total_transfer += _payment_amount(b, ["transfer", "transferencia"])
        tv = b.get("ecf_type") or "OTRO"
        by_tipo[tv] = by_tipo.get(tv, 0) + 1

    row = 6
    # Summary block 1: counts
    ws.cell(row=row, column=1, value="📊 RESUMEN DE FACTURAS").font = Font(bold=True, size=12, color="1F4E78")
    row += 1
    rows_data = [
        ("Total facturas emitidas", total_count, None),
        ("✓ Aprobadas por DGII", by_estado["Aprobada"], "Aprobada"),
        ("✗ Rechazadas por DGII", by_estado["Rechazada"], "Rechazada"),
        ("⚠ En Contingencia", by_estado["Contingencia"], "Contingencia"),
        ("📱 Cont. Manual (Plataforma externa)", by_estado["Cont. Manual"], "Cont. Manual"),
        ("⏳ Procesando", by_estado["Procesando"], "Procesando"),
        ("· Pendientes", by_estado["Pendiente"], "Pendiente"),
    ]
    for label, value, estado_key in rows_data:
        c1 = ws.cell(row=row, column=1, value=label)
        c2 = ws.cell(row=row, column=2, value=value)
        c1.alignment = LEFT
        c2.alignment = RIGHT
        c2.font = Font(bold=True, size=11)
        if estado_key:
            c1.fill = PatternFill("solid", fgColor=ESTADO_COLORS[estado_key])
            c1.font = Font(color=ESTADO_TEXT_COLORS[estado_key], size=10, bold=True)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="💰 TOTALES FISCALES (RD$)").font = Font(bold=True, size=12, color="1F4E78")
    row += 1
    for label, val in [
        ("Total NETO", total_neto),
        ("Total ITBIS (18%)", total_itbis),
        ("Total Propina/LEY (10%)", total_ley),
        ("TOTAL VENTAS", total_total),
    ]:
        ws.cell(row=row, column=1, value=label).alignment = LEFT
        c = ws.cell(row=row, column=2, value=round(val, 2))
        c.number_format = '#,##0.00'
        c.alignment = RIGHT
        c.font = Font(bold=True)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="💳 FORMAS DE PAGO (RD$)").font = Font(bold=True, size=12, color="1F4E78")
    row += 1
    for label, val in [
        ("Efectivo", total_efectivo),
        ("Tarjeta Visa", total_visa),
        ("Tarjeta MasterCard", total_mc),
        ("American Express (AMEX)", total_amex),
        ("Transferencia", total_transfer),
    ]:
        ws.cell(row=row, column=1, value=label).alignment = LEFT
        c = ws.cell(row=row, column=2, value=round(val, 2))
        c.number_format = '#,##0.00'
        c.alignment = RIGHT
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="📋 POR TIPO DE VENTA").font = Font(bold=True, size=12, color="1F4E78")
    row += 1
    for tipo, count in sorted(by_tipo.items()):
        ws.cell(row=row, column=1, value=tipo).alignment = LEFT
        ws.cell(row=row, column=2, value=count).alignment = RIGHT
        row += 1

    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 20

    # ── Sheet 2: DETALLE
    ws2 = wb.create_sheet("DETALLE")
    _write_title_block(ws2, title, subtitle, period_label, business["name"])

    headers = ["# Factura", "e-NCF", "Tipo Venta", "Nombre Comercial", "RNC",
               "Neto", "ITBIS", "LEY 10%", "Total",
               "Efectivo", "Tarjeta Visa", "Tarjeta MC", "AMEX", "Transferencia",
               "Fecha", "ESTADO"]
    if include_operational:
        headers += ["Cajero", "Mesero", "Mesa", "Hora"]

    _apply_header(ws2, 5, headers)

    row = 6
    for b in bills:
        estado = _classify_estado(b)
        paid_at = b.get("paid_at") or b.get("created_at") or ""
        fecha_str = ""
        hora_str = ""
        if paid_at:
            try:
                # paid_at is UTC ISO; convert to RD tz for display
                from zoneinfo import ZoneInfo
                dt = datetime.fromisoformat(str(paid_at).replace("Z", "+00:00"))
                local = dt.astimezone(ZoneInfo("America/Santo_Domingo"))
                fecha_str = local.strftime("%d/%m/%Y")
                hora_str = local.strftime("%H:%M:%S")
            except Exception:
                fecha_str = str(paid_at)[:10]

        cust = b.get("customer") or {}
        rnc = cust.get("rnc") or b.get("fiscal_id") or b.get("rnc") or ""
        razon = cust.get("razon_social") or b.get("razon_social") or cust.get("name") or ""

        values = [
            b.get("transaction_number") or b.get("fiscal_id") or "",
            b.get("ecf_encf") or b.get("ncf") or "",
            b.get("ecf_type") or "E32",
            razon,
            rnc,
            round(float(b.get("subtotal") or 0), 2),
            round(float(b.get("itbis") or 0), 2),
            round(float(b.get("tip") or b.get("service_charge") or 0), 2),
            round(float(b.get("total") or 0), 2),
            _payment_amount(b, ["efectivo", "cash"]),
            _payment_amount(b, ["visa"]),
            _payment_amount(b, ["master", "mc "]),
            _payment_amount(b, ["amex", "american"]),
            _payment_amount(b, ["transfer", "transferencia"]),
            fecha_str,
            estado,
        ]
        if include_operational:
            values += [
                b.get("cashier_name") or "",
                b.get("waiter_name") or "",
                b.get("table_number") or "",
                hora_str,
            ]

        for col_idx, val in enumerate(values, start=1):
            c = ws2.cell(row=row, column=col_idx, value=val)
            c.border = BORDER
            if col_idx in (1,):
                c.alignment = CENTER
            elif col_idx in (6, 7, 8, 9, 10, 11, 12, 13, 14):
                c.alignment = RIGHT
                c.number_format = '#,##0.00'
            elif col_idx == 16:
                _paint_estado_cell(c, estado)
            else:
                c.alignment = LEFT
        row += 1

    # Totals row
    if bills:
        tot_cell = ws2.cell(row=row, column=1, value="TOTAL")
        tot_cell.fill = TOTAL_FILL
        tot_cell.font = TOTAL_FONT
        tot_cell.alignment = CENTER
        for col_idx in range(2, len(headers) + 1):
            cell = ws2.cell(row=row, column=col_idx)
            cell.fill = TOTAL_FILL
            cell.font = TOTAL_FONT
            if col_idx in (6, 7, 8, 9, 10, 11, 12, 13, 14):
                col_letter = get_column_letter(col_idx)
                cell.value = f"=SUM({col_letter}6:{col_letter}{row - 1})"
                cell.number_format = '#,##0.00'
                cell.alignment = RIGHT

    ws2.freeze_panes = "A6"
    _autosize(ws2)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _build_compras_workbook(expenses: list, period_label: str, business: dict) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "RESUMEN"
    _write_title_block(ws, "REPORTE 606 — COMPRAS", "Gastos y compras a proveedores", period_label, business["name"])

    total_count = len(expenses)
    total_neto = total_itbis = total_total = 0.0
    by_supplier = {}
    for e in expenses:
        total_neto += float(e.get("subtotal") or 0)
        total_itbis += float(e.get("itbis") or 0)
        total_total += float(e.get("total") or 0)
        s = e.get("supplier_name") or "Sin proveedor"
        by_supplier[s] = by_supplier.get(s, 0) + float(e.get("total") or 0)

    row = 6
    ws.cell(row=row, column=1, value="📊 RESUMEN").font = Font(bold=True, size=12, color="1F4E78")
    row += 1
    for label, val, fmt in [
        ("Total compras registradas", total_count, "#,##0"),
        ("Total NETO (RD$)", round(total_neto, 2), "#,##0.00"),
        ("Total ITBIS pagado (RD$)", round(total_itbis, 2), "#,##0.00"),
        ("TOTAL GENERAL (RD$)", round(total_total, 2), "#,##0.00"),
    ]:
        ws.cell(row=row, column=1, value=label).alignment = LEFT
        c = ws.cell(row=row, column=2, value=val)
        c.alignment = RIGHT
        c.number_format = fmt
        c.font = Font(bold=True)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="🏢 POR PROVEEDOR").font = Font(bold=True, size=12, color="1F4E78")
    row += 1
    for name, amt in sorted(by_supplier.items(), key=lambda x: -x[1]):
        ws.cell(row=row, column=1, value=name).alignment = LEFT
        c = ws.cell(row=row, column=2, value=round(amt, 2))
        c.alignment = RIGHT
        c.number_format = '#,##0.00'
        row += 1

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 20

    # DETALLE
    ws2 = wb.create_sheet("DETALLE")
    _write_title_block(ws2, "REPORTE 606 — DETALLE", "Compras y gastos", period_label, business["name"])
    headers = ["# Gasto", "NCF Proveedor", "Tipo NCF", "Proveedor", "RNC Proveedor",
               "Neto", "ITBIS", "Total", "Forma Pago", "Categoría", "Fecha", "Referencia"]
    _apply_header(ws2, 5, headers)
    row = 6
    for e in expenses:
        fecha = e.get("expense_date") or ""
        fecha_str = str(fecha)[:10]
        values = [
            e.get("expense_number") or e.get("id", "")[:8],
            e.get("ncf") or "",
            e.get("ncf_type") or "",
            e.get("supplier_name") or "",
            e.get("supplier_rnc") or "",
            round(float(e.get("subtotal") or 0), 2),
            round(float(e.get("itbis") or 0), 2),
            round(float(e.get("total") or 0), 2),
            e.get("payment_method") or "",
            e.get("category") or "",
            fecha_str,
            e.get("reference") or "",
        ]
        for col_idx, val in enumerate(values, start=1):
            c = ws2.cell(row=row, column=col_idx, value=val)
            c.border = BORDER
            if col_idx in (6, 7, 8):
                c.alignment = RIGHT
                c.number_format = '#,##0.00'
            else:
                c.alignment = LEFT
        row += 1
    if expenses:
        tot_cell = ws2.cell(row=row, column=1, value="TOTAL")
        tot_cell.fill = TOTAL_FILL
        tot_cell.font = TOTAL_FONT
        tot_cell.alignment = CENTER
        for col_idx in range(2, len(headers) + 1):
            cell = ws2.cell(row=row, column=col_idx)
            cell.fill = TOTAL_FILL
            cell.font = TOTAL_FONT
            if col_idx in (6, 7, 8):
                col_letter = get_column_letter(col_idx)
                cell.value = f"=SUM({col_letter}6:{col_letter}{row - 1})"
                cell.number_format = '#,##0.00'
                cell.alignment = RIGHT
    ws2.freeze_panes = "A6"
    _autosize(ws2)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ═══════════════════════════════════════════════════════════════
# Endpoints
# ═══════════════════════════════════════════════════════════════

def _period_label(date_from: str, date_to: str) -> str:
    return f"Del {date_from} al {date_to}"


def _xlsx_response(buf: BytesIO, filename: str) -> StreamingResponse:
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/607")
async def report_607(
    date_from: str = Query(..., description="YYYY-MM-DD"),
    date_to: str = Query(..., description="YYYY-MM-DD"),
    user=Depends(get_current_user),
):
    """Reporte 607 — Ventas a DGII (formato detallado)."""
    bills = await _fetch_sales_bills(date_from, date_to)
    # 607 covers ALL sales (consumidor final + crédito fiscal + gubernamental)
    bills_607 = [b for b in bills if b.get("paid_at")]
    business = await _get_business_info()
    buf = _build_ventas_workbook(
        title="REPORTE 607 — VENTAS DGII",
        subtitle="Detalle de ventas fiscales",
        bills=bills_607,
        period_label=_period_label(date_from, date_to),
        business=business,
    )
    fname = f"607_{date_from}_al_{date_to}.xlsx"
    return _xlsx_response(buf, fname)


@router.get("/606")
async def report_606(
    date_from: str = Query(...),
    date_to: str = Query(...),
    user=Depends(get_current_user),
):
    """Reporte 606 — Compras DGII (formato detallado)."""
    expenses = await _fetch_expenses(date_from, date_to)
    business = await _get_business_info()
    buf = _build_compras_workbook(expenses, _period_label(date_from, date_to), business)
    fname = f"606_{date_from}_al_{date_to}.xlsx"
    return _xlsx_response(buf, fname)


@router.get("/608")
async def report_608(
    date_from: str = Query(...),
    date_to: str = Query(...),
    user=Depends(get_current_user),
):
    """Reporte 608 — Anulaciones / Notas de Crédito (formato detallado)."""
    bills = await _fetch_sales_bills(date_from, date_to)
    notes = await _fetch_credit_notes(date_from, date_to)
    business = await _get_business_info()

    # Build "virtual bills" from credit_notes + REJECTED bills
    virtual = []
    for n in notes:
        virtual.append({
            "transaction_number": n.get("transaction_number") or n.get("id", "")[:8],
            "ecf_encf": n.get("encf") or n.get("ncf") or "",
            "ecf_type": "E34",
            "customer": {"rnc": n.get("customer_rnc") or "", "razon_social": n.get("customer_name") or ""},
            "subtotal": n.get("subtotal") or 0,
            "itbis": n.get("itbis") or 0,
            "tip": 0,
            "total": n.get("total") or 0,
            "payments": [],
            "paid_at": n.get("created_at"),
            "ecf_status": n.get("ecf_status") or "FINISHED",
            "cashier_name": n.get("created_by_name") or "",
            "ecf_reject_reason": n.get("ecf_reject_reason"),
        })
    rejected = [b for b in bills if (b.get("ecf_status") or "").upper() == "REJECTED"]
    combined = virtual + rejected

    buf = _build_ventas_workbook(
        title="REPORTE 608 — ANULACIONES",
        subtitle="Notas de Crédito (E34) + Facturas Rechazadas",
        bills=combined,
        period_label=_period_label(date_from, date_to),
        business=business,
    )
    fname = f"608_{date_from}_al_{date_to}.xlsx"
    return _xlsx_response(buf, fname)


@router.get("/ventas-generales")
async def report_ventas_generales(
    date_from: str = Query(...),
    date_to: str = Query(...),
    user=Depends(get_current_user),
):
    """Reporte General de Ventas — Todas las facturas del período con detalles operacionales."""
    bills = await _fetch_sales_bills(date_from, date_to)
    business = await _get_business_info()
    buf = _build_ventas_workbook(
        title="REPORTE GENERAL DE VENTAS",
        subtitle="Todas las facturas del período",
        bills=bills,
        period_label=_period_label(date_from, date_to),
        business=business,
        include_operational=True,
    )
    fname = f"ventas_generales_{date_from}_al_{date_to}.xlsx"
    return _xlsx_response(buf, fname)


# ═══════════════════════════════════════════════════════════════
# Ventas por Categoría — Hierarchical report (HTML/PDF/Excel)
# ═══════════════════════════════════════════════════════════════

async def _fetch_category_breakdown(date_from: str, date_to: str) -> list:
    """Fetch category → products breakdown directly (avoids FastAPI endpoint invocation issues)."""
    bills = await db.bills.find(
        {"status": "paid", "training_mode": {"$ne": True}}, {"_id": 0}
    ).to_list(10000)
    filtered = [b for b in bills if date_from <= b.get("business_date", (b.get("paid_at") or "")[:10]) <= date_to]

    products = await db.products.find({}, {"_id": 0}).to_list(5000)
    prod_map = {p["id"]: p for p in products}
    prod_name_map = {p.get("name", "").lower(): p for p in products}
    cats_db = await db.categories.find({}, {"_id": 0}).to_list(200)
    cat_name_map = {c["id"]: c.get("name", c["id"]) for c in cats_db}

    categories = {}
    for bill in filtered:
        for item in bill.get("items", []):
            prod = prod_map.get(item.get("product_id"), prod_map.get(item.get("item_id"), {}))
            if not prod:
                prod = prod_name_map.get((item.get("product_name") or "").lower(), {})
            cat = prod.get("category_id", "sin_categoria")
            prod_name = item.get("product_name") or prod.get("name") or "Sin Nombre"
            if cat not in categories:
                categories[cat] = {"category": cat, "total": 0, "quantity": 0, "products": {}}
            item_total = item.get("total", item.get("subtotal", item.get("unit_price", 0) * item.get("quantity", 1)))
            item_qty = item.get("quantity", 1)
            categories[cat]["total"] += item_total
            categories[cat]["quantity"] += item_qty
            p_bucket = categories[cat]["products"].setdefault(prod_name, {"name": prod_name, "total": 0, "quantity": 0})
            p_bucket["total"] += item_total
            p_bucket["quantity"] += item_qty

    result = []
    for cat_id, data in categories.items():
        data["category"] = cat_name_map.get(cat_id, cat_id if len(str(cat_id)) < 30 else "Sin Categoría")
        data["total"] = round(data["total"], 2)
        plist = [{"name": p["name"], "total": round(p["total"], 2), "quantity": p["quantity"]}
                 for p in data["products"].values()]
        plist.sort(key=lambda x: x["name"].upper())
        data["products"] = plist
        result.append(data)
    result.sort(key=lambda x: -x["total"])
    return result


def _build_category_xlsx(data: list, period_label: str, business: dict) -> BytesIO:
    """
    Build hierarchical Ventas por Categoría XLSX.
    Columns: Descripción | Total | Cantidad
    Layout: Category (bold) → Products (indent) → Subtotal (bold, SUM formula) → blank row
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Ventas por Categoría"

    # Header block (sober B/W)
    title_font = Font(bold=True, size=14, color="000000")
    subtitle_font = Font(size=10, color="555555")

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    c = ws.cell(row=1, column=1, value=business["name"])
    c.font = title_font; c.alignment = CENTER
    ws.row_dimensions[1].height = 22

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=3)
    c = ws.cell(row=2, column=1, value=f"RNC: {business['rnc']}")
    c.font = subtitle_font; c.alignment = CENTER

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=3)
    c = ws.cell(row=3, column=1, value=f"Ventas por Categoría · {period_label}")
    c.font = Font(bold=True, size=11); c.alignment = CENTER

    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=3)
    c = ws.cell(row=4, column=1, value=f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    c.font = subtitle_font; c.alignment = CENTER

    # Column headers (B/W: black fill, white text)
    headers = ["Descripción", "Total", "Cantidad"]
    bw_header_fill = PatternFill("solid", fgColor="111111")
    bw_header_font = Font(bold=True, color="FFFFFF", size=11)
    for idx, h in enumerate(headers, start=1):
        cc = ws.cell(row=6, column=idx, value=h)
        cc.fill = bw_header_fill; cc.font = bw_header_font
        cc.alignment = CENTER; cc.border = BORDER

    row = 7
    grand_total_cells = []  # rows containing category subtotals (for grand-total SUM)
    for cat in data:
        # Category name (bold)
        c_name = ws.cell(row=row, column=1, value=cat["category"])
        c_name.font = Font(bold=True, size=11)
        c_name.alignment = LEFT
        row += 1

        # Products
        product_first_row = row
        for p in cat.get("products", []):
            # Indent via cell alignment
            pc1 = ws.cell(row=row, column=1, value=p["name"])
            pc1.alignment = Alignment(horizontal="left", vertical="center", indent=2)
            pc2 = ws.cell(row=row, column=2, value=round(p["total"], 2))
            pc2.number_format = '#,##0.00'; pc2.alignment = RIGHT
            pc3 = ws.cell(row=row, column=3, value=p["quantity"])
            pc3.alignment = RIGHT
            row += 1
        product_last_row = row - 1

        # Subtotal row (same category name repeated, bold, with SUM formula)
        sub1 = ws.cell(row=row, column=1, value=cat["category"])
        sub1.font = Font(bold=True, size=11); sub1.alignment = LEFT
        sub1.border = Border(top=Side(style="thin", color="000000"))
        if product_last_row >= product_first_row:
            sub2 = ws.cell(row=row, column=2, value=f"=SUM(B{product_first_row}:B{product_last_row})")
            sub3 = ws.cell(row=row, column=3, value=f"=SUM(C{product_first_row}:C{product_last_row})")
        else:
            sub2 = ws.cell(row=row, column=2, value=0)
            sub3 = ws.cell(row=row, column=3, value=0)
        sub2.font = Font(bold=True); sub2.number_format = '#,##0.00'; sub2.alignment = RIGHT
        sub2.border = Border(top=Side(style="thin", color="000000"))
        sub3.font = Font(bold=True); sub3.alignment = RIGHT
        sub3.border = Border(top=Side(style="thin", color="000000"))
        grand_total_cells.append(row)
        row += 1

        # Blank spacer
        row += 1

    # Grand total row
    gt1 = ws.cell(row=row, column=1, value="TOTAL GENERAL")
    gt1.font = Font(bold=True, size=12)
    gt1.border = Border(top=Side(style="medium", color="000000"))
    gt1.alignment = LEFT
    if grand_total_cells:
        cells_total = ",".join([f"B{r}" for r in grand_total_cells])
        cells_qty = ",".join([f"C{r}" for r in grand_total_cells])
        gt2 = ws.cell(row=row, column=2, value=f"=SUM({cells_total})")
        gt3 = ws.cell(row=row, column=3, value=f"=SUM({cells_qty})")
    else:
        gt2 = ws.cell(row=row, column=2, value=0)
        gt3 = ws.cell(row=row, column=3, value=0)
    gt2.font = Font(bold=True, size=12); gt2.number_format = '#,##0.00'; gt2.alignment = RIGHT
    gt2.border = Border(top=Side(style="medium", color="000000"))
    gt3.font = Font(bold=True, size=12); gt3.alignment = RIGHT
    gt3.border = Border(top=Side(style="medium", color="000000"))

    # Footer row
    row += 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    footer = ws.cell(row=row, column=1,
                     value=f"Documento generado automáticamente por {business['name']} | {datetime.now().strftime('%d/%m/%Y %H:%M')} | Válido para auditoría interna")
    footer.font = Font(italic=True, size=9, color="555555")
    footer.alignment = CENTER

    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 12
    ws.freeze_panes = "A7"

    buf = BytesIO()
    wb.save(buf); buf.seek(0)
    return buf


def _build_category_html(data: list, period_label: str, business: dict, for_pdf: bool = False) -> str:
    """Build sober B/W HTML for PDF generation. for_pdf=True applies print-optimized styles."""
    rows = []
    grand_total = 0.0
    grand_qty = 0

    for cat in data:
        cat_total = sum(p["total"] for p in cat.get("products", []))
        cat_qty = sum(p["quantity"] for p in cat.get("products", []))
        grand_total += cat_total
        grand_qty += cat_qty
        rows.append(f'<tr class="cat-header"><td colspan="3"><strong>{cat["category"]}</strong></td></tr>')
        for p in cat.get("products", []):
            rows.append(
                f'<tr class="prod-row">'
                f'<td class="indent">{p["name"]}</td>'
                f'<td class="money">{p["total"]:,.2f}</td>'
                f'<td class="qty">{p["quantity"]}</td></tr>'
            )
        rows.append(
            f'<tr class="cat-subtotal">'
            f'<td><strong>{cat["category"]}</strong></td>'
            f'<td class="money"><strong>{cat_total:,.2f}</strong></td>'
            f'<td class="qty"><strong>{cat_qty}</strong></td></tr>'
        )
        rows.append('<tr class="spacer"><td colspan="3">&nbsp;</td></tr>')

    rows.append(
        f'<tr class="grand-total">'
        f'<td><strong>TOTAL GENERAL</strong></td>'
        f'<td class="money"><strong>{grand_total:,.2f}</strong></td>'
        f'<td class="qty"><strong>{grand_qty}</strong></td></tr>'
    )

    return f"""
    <!DOCTYPE html>
    <html lang="es">
    <head>
      <meta charset="utf-8" />
      <title>Ventas por Categoría</title>
      <style>
        @page {{ size: Letter; margin: 18mm 14mm; }}
        * {{ box-sizing: border-box; }}
        body {{ font-family: 'Helvetica', 'Arial', sans-serif; color: #000; font-size: 11pt; margin: 0; padding: 0; }}
        .header {{ text-align: center; border-bottom: 2px solid #000; padding-bottom: 8px; margin-bottom: 12px; }}
        .header h1 {{ margin: 0; font-size: 16pt; color: #000; }}
        .header .rnc {{ color: #444; font-size: 10pt; margin-top: 4px; }}
        .header .title {{ font-size: 12pt; font-weight: bold; margin-top: 6px; }}
        .header .date {{ color: #666; font-size: 9pt; margin-top: 2px; }}
        table {{ width: 100%; border-collapse: collapse; margin-top: 8px; }}
        thead th {{ background: #111; color: #fff; padding: 6px 8px; text-align: left; font-size: 10pt; }}
        thead th.money, thead th.qty {{ text-align: right; }}
        td {{ padding: 4px 8px; font-size: 10pt; vertical-align: middle; }}
        td.money {{ text-align: right; font-variant-numeric: tabular-nums; }}
        td.qty {{ text-align: right; font-variant-numeric: tabular-nums; width: 80px; }}
        td.indent {{ padding-left: 24px; }}
        tr.cat-header td {{ background: transparent; border-top: 1px solid #000; padding-top: 6px; }}
        tr.cat-subtotal td {{ border-top: 1px solid #000; padding-top: 4px; }}
        tr.spacer td {{ padding: 2px; }}
        tr.grand-total td {{ border-top: 2px solid #000; border-bottom: 2px solid #000; padding: 8px; background: #f5f5f5; font-size: 11pt; }}
        .footer {{ margin-top: 24px; text-align: center; color: #666; font-size: 8pt; font-style: italic; }}
        @media print {{ body {{ background: white; }} tr.grand-total td {{ background: #fff; }} }}
      </style>
    </head>
    <body>
      <div class="header">
        <h1>{business['name']}</h1>
        <div class="rnc">RNC: {business['rnc']}</div>
        <div class="title">Ventas por Categoría</div>
        <div class="date">{period_label} · Generado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}</div>
      </div>
      <table>
        <thead>
          <tr>
            <th>Descripción</th>
            <th class="money">Total</th>
            <th class="qty">Cantidad</th>
          </tr>
        </thead>
        <tbody>
          {''.join(rows)}
        </tbody>
      </table>
      <div class="footer">
        Documento generado automáticamente por {business['name']} | {datetime.now().strftime('%d/%m/%Y %H:%M')} | Válido para auditoría interna
      </div>
    </body>
    </html>
    """


@router.get("/ventas-por-categoria/xlsx")
async def ventas_por_categoria_xlsx(
    date_from: str = Query(...),
    date_to: str = Query(...),
    user=Depends(get_current_user),
):
    data = await _fetch_category_breakdown(date_from, date_to)
    business = await _get_business_info()
    buf = _build_category_xlsx(data, _period_label(date_from, date_to), business)
    fname = f"VentasPorCategoria_{date_from}_al_{date_to}.xlsx"
    return _xlsx_response(buf, fname)


@router.get("/ventas-por-categoria/pdf")
async def ventas_por_categoria_pdf(
    date_from: str = Query(...),
    date_to: str = Query(...),
    user=Depends(get_current_user),
):
    """Server-side PDF generation via WeasyPrint (stable, pure-Python, no Chromium needed)."""
    try:
        from weasyprint import HTML
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"WeasyPrint no disponible: {e}")

    data = await _fetch_category_breakdown(date_from, date_to)
    business = await _get_business_info()
    html = _build_category_html(data, _period_label(date_from, date_to), business, for_pdf=True)
    pdf_bytes = HTML(string=html).write_pdf()
    fname = f"VentasPorCategoria_{date_from}_al_{date_to}.pdf"
    return StreamingResponse(
        BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )



# ═══════════════════════════════════════════════════════════════
# CIERRE DE CAJA — Hierarchical (Empleado → Turno → Método → Trans)
# ═══════════════════════════════════════════════════════════════

async def _fetch_cash_close_tree(date_from: str, date_to: str) -> dict:
    """Reuse the pure hierarchical helper from reports router."""
    from routers.reports import _cash_close_hierarchical_impl  # local import to avoid circular at module load
    return await _cash_close_hierarchical_impl(date_from=date_from, date_to=date_to)


def _fmt_dt(s: str) -> str:
    """Format ISO datetime to 'DD/MM/YYYY HH:MM' (fallbacks to raw value)."""
    if not s:
        return "-"
    try:
        if "T" in s:
            base = s.split(".")[0].replace("Z", "")
            return datetime.fromisoformat(base).strftime("%d/%m/%Y %I:%M %p")
        return s
    except Exception:
        return s


def _build_cash_close_xlsx(tree: dict, period_label: str, business: dict, detailed: bool) -> BytesIO:
    """Build XLSX with Employee → Shift → Payment Method (→ Transactions if detailed).
    Uses openpyxl row outline grouping and real SUM() formulas for subtotals/totals.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Cierre de Caja"

    BLACK_FILL = PatternFill("solid", fgColor="111111")
    WHITE_FONT = Font(bold=True, color="FFFFFF", size=11)
    BOLD = Font(bold=True)
    THIN_BLACK = Side(style="thin", color="000000")

    # Title block
    ws.cell(row=1, column=1, value=business.get("name", "VexlyPOS")).font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"RNC: {business.get('rnc','')}").font = Font(size=10, color="555555")
    ws.cell(row=3, column=1, value=f"Cierre de Caja — {'Detallado' if detailed else 'Resumido'}").font = Font(bold=True, size=12)
    ws.cell(row=4, column=1, value=period_label).font = Font(size=10, color="555555")

    # Header row (row 6)
    headers = ["Descripción", "Shift Start", "Shift End", "Trans #", "Total", "Propinas", "Total + Propinas"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=6, column=i, value=h)
        c.font = WHITE_FONT
        c.fill = BLACK_FILL
        c.alignment = CENTER if i >= 4 else LEFT
        c.border = Border(top=THIN_BLACK, bottom=THIN_BLACK, left=THIN_BLACK, right=THIN_BLACK)

    row = 7
    employee_total_rows = []

    for emp in tree.get("employees", []):
        # Employee header (outlineLevel 0)
        ws.cell(row=row, column=1, value=emp["name"]).font = Font(bold=True, size=11)
        row += 1

        shift_total_rows_for_emp = []

        for sh in emp.get("shifts", []):
            # Shift header (outlineLevel 1)
            c = ws.cell(row=row, column=1, value="   Turno")
            c.font = BOLD
            ws.cell(row=row, column=2, value=_fmt_dt(sh.get("shift_start"))).font = Font(size=10)
            ws.cell(row=row, column=3, value=_fmt_dt(sh.get("shift_end"))).font = Font(size=10)
            ws.row_dimensions[row].outlineLevel = 1
            row += 1

            pm_subtotal_rows_for_shift = []

            for pm in sh.get("payment_methods", []):
                # Payment method header (outlineLevel 2)
                ws.cell(row=row, column=1, value=f"      {pm['name']}").font = BOLD
                ws.row_dimensions[row].outlineLevel = 2
                row += 1

                tx_rows = []
                if detailed:
                    for tx in pm.get("transactions", []):
                        ws.cell(row=row, column=1, value="         " + (tx.get("paid_at") or ""))
                        ws.cell(row=row, column=4, value=tx.get("trans_number")).alignment = CENTER
                        t_cell = ws.cell(row=row, column=5, value=float(tx.get("total") or 0))
                        p_cell = ws.cell(row=row, column=6, value=float(tx.get("tips") or 0))
                        tp_cell = ws.cell(row=row, column=7, value=float(tx.get("total_with_tips") or 0))
                        for cc in (t_cell, p_cell, tp_cell):
                            cc.number_format = '#,##0.00'
                            cc.alignment = RIGHT
                        ws.row_dimensions[row].outlineLevel = 3
                        tx_rows.append(row)
                        row += 1

                # Payment method subtotal
                sub = pm.get("subtotal", {})
                c1 = ws.cell(row=row, column=1, value=f"      {pm['name']} [{sub.get('count', 0)}]")
                c1.font = BOLD
                if detailed and tx_rows:
                    ws.cell(row=row, column=5, value=f"=SUM(E{tx_rows[0]}:E{tx_rows[-1]})").number_format = '#,##0.00'
                    ws.cell(row=row, column=6, value=f"=SUM(F{tx_rows[0]}:F{tx_rows[-1]})").number_format = '#,##0.00'
                    ws.cell(row=row, column=7, value=f"=SUM(G{tx_rows[0]}:G{tx_rows[-1]})").number_format = '#,##0.00'
                else:
                    ws.cell(row=row, column=5, value=float(sub.get("total", 0) or 0)).number_format = '#,##0.00'
                    ws.cell(row=row, column=6, value=float(sub.get("tips", 0) or 0)).number_format = '#,##0.00'
                    ws.cell(row=row, column=7, value=float(sub.get("total_with_tips", 0) or 0)).number_format = '#,##0.00'
                for col in (5, 6, 7):
                    cc = ws.cell(row=row, column=col)
                    cc.font = BOLD
                    cc.alignment = RIGHT
                    cc.border = Border(top=THIN_BLACK)
                ws.row_dimensions[row].outlineLevel = 2
                pm_subtotal_rows_for_shift.append(row)
                row += 1

            # Shift totals
            sht = sh.get("shift_totals", {})
            c1 = ws.cell(row=row, column=1, value=f"   Shift Totals [{sht.get('count', 0)}]")
            c1.font = Font(bold=True, size=11)
            if pm_subtotal_rows_for_shift:
                cells_e = ",".join([f"E{r}" for r in pm_subtotal_rows_for_shift])
                cells_f = ",".join([f"F{r}" for r in pm_subtotal_rows_for_shift])
                cells_g = ",".join([f"G{r}" for r in pm_subtotal_rows_for_shift])
                ws.cell(row=row, column=5, value=f"=SUM({cells_e})").number_format = '#,##0.00'
                ws.cell(row=row, column=6, value=f"=SUM({cells_f})").number_format = '#,##0.00'
                ws.cell(row=row, column=7, value=f"=SUM({cells_g})").number_format = '#,##0.00'
            else:
                ws.cell(row=row, column=5, value=0).number_format = '#,##0.00'
                ws.cell(row=row, column=6, value=0).number_format = '#,##0.00'
                ws.cell(row=row, column=7, value=0).number_format = '#,##0.00'
            for col in (5, 6, 7):
                cc = ws.cell(row=row, column=col)
                cc.font = Font(bold=True)
                cc.alignment = RIGHT
                cc.border = Border(top=THIN_BLACK, bottom=THIN_BLACK)
            ws.row_dimensions[row].outlineLevel = 1
            shift_total_rows_for_emp.append(row)
            row += 1

        # Employee total
        et = emp.get("employee_totals", {})
        c1 = ws.cell(row=row, column=1, value=f"TOTAL EMPLEADO [{et.get('count', 0)}]")
        c1.font = Font(bold=True, size=11)
        if shift_total_rows_for_emp:
            cells_e = ",".join([f"E{r}" for r in shift_total_rows_for_emp])
            cells_f = ",".join([f"F{r}" for r in shift_total_rows_for_emp])
            cells_g = ",".join([f"G{r}" for r in shift_total_rows_for_emp])
            ws.cell(row=row, column=5, value=f"=SUM({cells_e})").number_format = '#,##0.00'
            ws.cell(row=row, column=6, value=f"=SUM({cells_f})").number_format = '#,##0.00'
            ws.cell(row=row, column=7, value=f"=SUM({cells_g})").number_format = '#,##0.00'
        else:
            ws.cell(row=row, column=5, value=0).number_format = '#,##0.00'
            ws.cell(row=row, column=6, value=0).number_format = '#,##0.00'
            ws.cell(row=row, column=7, value=0).number_format = '#,##0.00'
        for col in (5, 6, 7):
            cc = ws.cell(row=row, column=col)
            cc.font = Font(bold=True)
            cc.alignment = RIGHT
            cc.border = Border(top=Side(style="medium", color="000000"), bottom=THIN_BLACK)
        employee_total_rows.append(row)
        row += 2  # spacer

    # Grand total
    gt = tree.get("grand_totals", {})
    c1 = ws.cell(row=row, column=1, value=f"TOTAL GENERAL [{gt.get('count', 0)}]")
    c1.font = Font(bold=True, size=12)
    if employee_total_rows:
        cells_e = ",".join([f"E{r}" for r in employee_total_rows])
        cells_f = ",".join([f"F{r}" for r in employee_total_rows])
        cells_g = ",".join([f"G{r}" for r in employee_total_rows])
        ws.cell(row=row, column=5, value=f"=SUM({cells_e})").number_format = '#,##0.00'
        ws.cell(row=row, column=6, value=f"=SUM({cells_f})").number_format = '#,##0.00'
        ws.cell(row=row, column=7, value=f"=SUM({cells_g})").number_format = '#,##0.00'
    else:
        for col, v in ((5, 0), (6, 0), (7, 0)):
            ws.cell(row=row, column=col, value=v).number_format = '#,##0.00'
    for col in (5, 6, 7):
        cc = ws.cell(row=row, column=col)
        cc.font = Font(bold=True, size=12)
        cc.alignment = RIGHT
        cc.border = Border(top=Side(style="medium", color="000000"), bottom=Side(style="medium", color="000000"))

    # Footer
    row += 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    footer = ws.cell(
        row=row, column=1,
        value=f"Documento generado por {business.get('name','')} | {datetime.now().strftime('%d/%m/%Y %H:%M')} | Uso interno"
    )
    footer.font = Font(italic=True, size=9, color="555555")
    footer.alignment = CENTER

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 16
    ws.freeze_panes = "A7"
    ws.sheet_properties.outlinePr.summaryBelow = True

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _build_cash_close_html(tree: dict, period_label: str, business: dict, detailed: bool) -> str:
    """Print-friendly B/W HTML for PDF generation."""
    rows = []
    for emp in tree.get("employees", []):
        rows.append(
            f'<tr class="emp-header"><td colspan="7"><strong>{emp["name"]}</strong></td></tr>'
        )
        for sh in emp.get("shifts", []):
            rows.append(
                f'<tr class="shift-header">'
                f'<td class="ind1"><em>Turno</em></td>'
                f'<td>{_fmt_dt(sh.get("shift_start"))}</td>'
                f'<td>{_fmt_dt(sh.get("shift_end"))}</td>'
                f'<td colspan="4"></td></tr>'
            )
            for pm in sh.get("payment_methods", []):
                rows.append(
                    f'<tr class="pm-header"><td class="ind2"><strong>{pm["name"]}</strong></td>'
                    f'<td colspan="6"></td></tr>'
                )
                if detailed:
                    for tx in pm.get("transactions", []):
                        rows.append(
                            f'<tr class="tx">'
                            f'<td class="ind3">{_fmt_dt(tx.get("paid_at"))}</td>'
                            f'<td></td><td></td>'
                            f'<td class="center">{tx.get("trans_number", "")}</td>'
                            f'<td class="money">{float(tx.get("total") or 0):,.2f}</td>'
                            f'<td class="money">{float(tx.get("tips") or 0):,.2f}</td>'
                            f'<td class="money">{float(tx.get("total_with_tips") or 0):,.2f}</td></tr>'
                        )
                sub = pm.get("subtotal", {})
                rows.append(
                    f'<tr class="pm-sub">'
                    f'<td class="ind2"><strong>{pm["name"]} [{sub.get("count", 0)}]</strong></td>'
                    f'<td colspan="3"></td>'
                    f'<td class="money"><strong>{float(sub.get("total") or 0):,.2f}</strong></td>'
                    f'<td class="money"><strong>{float(sub.get("tips") or 0):,.2f}</strong></td>'
                    f'<td class="money"><strong>{float(sub.get("total_with_tips") or 0):,.2f}</strong></td></tr>'
                )
            sht = sh.get("shift_totals", {})
            rows.append(
                f'<tr class="shift-sub">'
                f'<td class="ind1"><strong>Shift Totals [{sht.get("count", 0)}]</strong></td>'
                f'<td colspan="3"></td>'
                f'<td class="money"><strong>{float(sht.get("total") or 0):,.2f}</strong></td>'
                f'<td class="money"><strong>{float(sht.get("tips") or 0):,.2f}</strong></td>'
                f'<td class="money"><strong>{float(sht.get("total_with_tips") or 0):,.2f}</strong></td></tr>'
            )
        et = emp.get("employee_totals", {})
        rows.append(
            f'<tr class="emp-total">'
            f'<td><strong>TOTAL EMPLEADO [{et.get("count", 0)}]</strong></td>'
            f'<td colspan="3"></td>'
            f'<td class="money"><strong>{float(et.get("total") or 0):,.2f}</strong></td>'
            f'<td class="money"><strong>{float(et.get("tips") or 0):,.2f}</strong></td>'
            f'<td class="money"><strong>{float(et.get("total_with_tips") or 0):,.2f}</strong></td></tr>'
        )
        rows.append('<tr class="spacer"><td colspan="7">&nbsp;</td></tr>')

    gt = tree.get("grand_totals", {})
    rows.append(
        f'<tr class="grand-total">'
        f'<td><strong>TOTAL GENERAL [{gt.get("count", 0)}]</strong></td>'
        f'<td colspan="3"></td>'
        f'<td class="money"><strong>{float(gt.get("total") or 0):,.2f}</strong></td>'
        f'<td class="money"><strong>{float(gt.get("tips") or 0):,.2f}</strong></td>'
        f'<td class="money"><strong>{float(gt.get("total_with_tips") or 0):,.2f}</strong></td></tr>'
    )

    view_label = "Detallado" if detailed else "Resumido"
    return f"""
    <!DOCTYPE html>
    <html lang="es">
    <head>
      <meta charset="utf-8" />
      <title>Cierre de Caja — {view_label}</title>
      <style>
        @page {{ size: Letter; margin: 16mm 12mm; }}
        * {{ box-sizing: border-box; }}
        body {{ font-family: 'Helvetica', 'Arial', sans-serif; color: #000; font-size: 10pt; margin: 0; padding: 0; }}
        .header {{ text-align: center; border-bottom: 2px solid #000; padding-bottom: 8px; margin-bottom: 10px; }}
        .header h1 {{ margin: 0; font-size: 15pt; color: #000; }}
        .header .rnc {{ color: #444; font-size: 9pt; margin-top: 2px; }}
        .header .title {{ font-size: 11pt; font-weight: bold; margin-top: 4px; }}
        .header .date {{ color: #666; font-size: 9pt; margin-top: 2px; }}
        table {{ width: 100%; border-collapse: collapse; margin-top: 6px; }}
        thead th {{ background: #111; color: #fff; padding: 5px 6px; font-size: 9pt; }}
        thead th.money {{ text-align: right; }}
        td {{ padding: 3px 6px; font-size: 9pt; vertical-align: middle; }}
        td.money {{ text-align: right; font-variant-numeric: tabular-nums; }}
        td.center {{ text-align: center; }}
        td.ind1 {{ padding-left: 14px; }}
        td.ind2 {{ padding-left: 28px; }}
        td.ind3 {{ padding-left: 42px; color: #333; }}
        tr.emp-header td {{ border-top: 1.5px solid #000; padding-top: 6px; font-size: 10pt; }}
        tr.shift-header td {{ border-top: 1px solid #888; }}
        tr.pm-sub td {{ border-top: 1px solid #000; }}
        tr.shift-sub td {{ border-top: 1px solid #000; border-bottom: 1px solid #000; }}
        tr.emp-total td {{ border-top: 1.5px solid #000; border-bottom: 1.5px solid #000; padding: 5px 6px; }}
        tr.spacer td {{ padding: 3px; }}
        tr.grand-total td {{ border-top: 2px solid #000; border-bottom: 2px solid #000; padding: 7px 6px; font-size: 11pt; }}
        .footer {{ margin-top: 16px; text-align: center; color: #555; font-size: 8pt; font-style: italic; }}
        @media print {{ body {{ background: white; }} thead {{ display: table-header-group; }} tr {{ page-break-inside: avoid; }} }}
      </style>
    </head>
    <body>
      <div class="header">
        <h1>{business['name']}</h1>
        <div class="rnc">RNC: {business['rnc']}</div>
        <div class="title">Cierre de Caja — {view_label}</div>
        <div class="date">{period_label} · Generado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}</div>
      </div>
      <table>
        <thead>
          <tr>
            <th>Descripción</th>
            <th>Shift Start</th>
            <th>Shift End</th>
            <th>Trans #</th>
            <th class="money">Total</th>
            <th class="money">Propinas</th>
            <th class="money">Total + Prop.</th>
          </tr>
        </thead>
        <tbody>
          {''.join(rows)}
        </tbody>
      </table>
      <div class="footer">
        Documento generado automáticamente por {business['name']} | {datetime.now().strftime('%d/%m/%Y %H:%M')} | Uso interno
      </div>
    </body>
    </html>
    """


@router.get("/cierre-caja/xlsx")
async def cierre_caja_xlsx(
    date_from: str = Query(...),
    date_to: str = Query(...),
    view: str = Query("resumida", regex="^(resumida|detallada)$"),
    user=Depends(get_current_user),
):
    detailed = (view == "detallada")
    tree = await _fetch_cash_close_tree(date_from, date_to)
    business = await _get_business_info()
    buf = _build_cash_close_xlsx(tree, _period_label(date_from, date_to), business, detailed)
    label = "Detallada" if detailed else "Resumida"
    fname = f"CierreCaja_{label}_{date_from}_al_{date_to}.xlsx"
    return _xlsx_response(buf, fname)


@router.get("/cierre-caja/pdf")
async def cierre_caja_pdf(
    date_from: str = Query(...),
    date_to: str = Query(...),
    view: str = Query("resumida", regex="^(resumida|detallada)$"),
    user=Depends(get_current_user),
):
    try:
        from weasyprint import HTML
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"WeasyPrint no disponible: {e}")
    detailed = (view == "detallada")
    tree = await _fetch_cash_close_tree(date_from, date_to)
    business = await _get_business_info()
    html = _build_cash_close_html(tree, _period_label(date_from, date_to), business, detailed)
    pdf_bytes = HTML(string=html).write_pdf()
    label = "Detallada" if detailed else "Resumida"
    fname = f"CierreCaja_{label}_{date_from}_al_{date_to}.pdf"
    return StreamingResponse(
        BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ═══════════════════════════════════════════════════════════════
# VENTAS POR HORA — PDF + XLSX (consume /reports/hourly-sales)
# ═══════════════════════════════════════════════════════════════

async def _fetch_hourly_data(date_from: str, date_to: str) -> list:
    """Aggregate paid bills by hour (0..23) over [date_from, date_to]. Returns list of 24 rows."""
    bills = await db.bills.find(
        {"status": "paid", "training_mode": {"$ne": True}}, {"_id": 0}
    ).to_list(20000)
    filtered = [
        b for b in bills
        if date_from <= (b.get("business_date") or (b.get("paid_at") or "")[:10]) <= date_to
    ]
    hourly = {f"{h:02d}": {"hour": f"{h:02d}:00", "total": 0.0, "bills": 0} for h in range(24)}
    for b in filtered:
        paid_at = b.get("paid_at") or ""
        if "T" in paid_at:
            hh = paid_at.split("T")[1][:2]
            if hh in hourly:
                hourly[hh]["total"] += float(b.get("total") or 0)
                hourly[hh]["bills"] += 1
    return list(hourly.values())


def _build_hourly_xlsx(rows: list, period_label: str, business: dict) -> BytesIO:
    """XLSX with Hour range | Bills | Total | Ticket Avg | % | with real SUM formulas on totals row."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Ventas por Hora"

    BLACK_FILL = PatternFill("solid", fgColor="111111")
    WHITE_FONT = Font(bold=True, color="FFFFFF", size=11)
    THIN_BLACK = Side(style="thin", color="000000")

    # Title block
    ws.cell(row=1, column=1, value=business.get("name", "VexlyPOS")).font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"RNC: {business.get('rnc','')}").font = Font(size=10, color="555555")
    ws.cell(row=3, column=1, value="Ventas por Hora").font = Font(bold=True, size=12)
    ws.cell(row=4, column=1, value=period_label).font = Font(size=10, color="555555")

    # Header row 6
    headers = ["Hora", "# Facturas", "Total Ventas", "Ticket Promedio", "% del Total"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=6, column=i, value=h)
        c.font = WHITE_FONT
        c.fill = BLACK_FILL
        c.alignment = CENTER if i >= 2 else LEFT
        c.border = Border(top=THIN_BLACK, bottom=THIN_BLACK, left=THIN_BLACK, right=THIN_BLACK)

    row_num = 7
    data_start = row_num
    for r in rows:
        h = int((r.get("hour") or "00:00")[:2])
        nxt = (h + 1) % 24
        hour_range = f"{h:02d}:00-{nxt:02d}:00"
        bills = int(r.get("bills") or 0)
        total = float(r.get("total") or 0)

        ws.cell(row=row_num, column=1, value=hour_range).font = Font(name="Courier New", size=10)
        ws.cell(row=row_num, column=2, value=bills).alignment = RIGHT
        c3 = ws.cell(row=row_num, column=3, value=total)
        c3.number_format = '#,##0.00'
        c3.alignment = RIGHT
        # Ticket avg formula
        c4 = ws.cell(row=row_num, column=4, value=f"=IF(B{row_num}>0,C{row_num}/B{row_num},0)")
        c4.number_format = '#,##0.00'
        c4.alignment = RIGHT
        ws.cell(row=row_num, column=5)
        c5 = ws.cell(row=row_num, column=5)
        c5.number_format = '0.0%'
        c5.alignment = RIGHT
        row_num += 1
    data_end = row_num - 1

    # Remove the accidentally-duplicated empty cell creation on column 5 (the value is back-filled below)
    # (no-op: openpyxl cell() retrieval is idempotent)

    # Total row (with SUM formulas)
    total_row = row_num
    c1 = ws.cell(row=total_row, column=1, value="TOTAL GENERAL")
    c1.font = Font(bold=True, size=11)
    c1.border = Border(top=Side(style="medium", color="000000"), bottom=Side(style="medium", color="000000"))
    ws.cell(row=total_row, column=2, value=f"=SUM(B{data_start}:B{data_end})")
    ws.cell(row=total_row, column=3, value=f"=SUM(C{data_start}:C{data_end})")
    ws.cell(row=total_row, column=4, value=f"=IF(B{total_row}>0,C{total_row}/B{total_row},0)")
    ws.cell(row=total_row, column=5, value=1.0)
    for col, fmt in ((2, '#,##0'), (3, '#,##0.00'), (4, '#,##0.00'), (5, '0.0%')):
        cc = ws.cell(row=total_row, column=col)
        cc.font = Font(bold=True, size=11)
        cc.number_format = fmt
        cc.alignment = RIGHT
        cc.border = Border(top=Side(style="medium", color="000000"), bottom=Side(style="medium", color="000000"))

    # Back-fill % formulas for data rows (referencing total row)
    for rn in range(data_start, data_end + 1):
        ws.cell(row=rn, column=5, value=f"=IF(C{total_row}>0,C{rn}/C{total_row},0)")

    # Footer
    row_num = total_row + 2
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=5)
    footer = ws.cell(
        row=row_num, column=1,
        value=f"Documento generado por {business.get('name','')} | {datetime.now().strftime('%d/%m/%Y %H:%M')} | Uso interno"
    )
    footer.font = Font(italic=True, size=9, color="555555")
    footer.alignment = CENTER

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 14
    ws.freeze_panes = "A7"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _build_hourly_html(rows: list, period_label: str, business: dict) -> str:
    """Print-friendly B/W HTML for PDF generation."""
    grand_total = sum(float(r.get("total") or 0) for r in rows)
    grand_bills = sum(int(r.get("bills") or 0) for r in rows)
    peak = None
    for r in rows:
        if (r.get("bills") or 0) > 0 and (peak is None or r["total"] > peak["total"]):
            peak = r

    trs = []
    for r in rows:
        h = int((r.get("hour") or "00:00")[:2])
        nxt = (h + 1) % 24
        hour_range = f"{h:02d}:00-{nxt:02d}:00"
        bills = int(r.get("bills") or 0)
        total = float(r.get("total") or 0)
        avg = (total / bills) if bills else 0
        pct = (total / grand_total * 100) if grand_total > 0 else 0
        is_peak = peak is not None and r.get("hour") == peak.get("hour")
        cls = "row-peak" if is_peak else ("row-empty" if bills == 0 else "")
        trs.append(
            f'<tr class="{cls}">'
            f'<td class="mono">{hour_range}</td>'
            f'<td class="num">{bills}</td>'
            f'<td class="money">{total:,.2f}</td>'
            f'<td class="money">{avg:,.2f}</td>'
            f'<td class="num">{pct:.1f}%</td></tr>'
        )
    avg_total = (grand_total / grand_bills) if grand_bills else 0
    trs.append(
        f'<tr class="grand-total">'
        f'<td><strong>TOTAL GENERAL</strong></td>'
        f'<td class="num"><strong>{grand_bills}</strong></td>'
        f'<td class="money"><strong>{grand_total:,.2f}</strong></td>'
        f'<td class="money"><strong>{avg_total:,.2f}</strong></td>'
        f'<td class="num"><strong>100.0%</strong></td></tr>'
    )

    return f"""
    <!DOCTYPE html>
    <html lang="es">
    <head>
      <meta charset="utf-8" />
      <title>Ventas por Hora</title>
      <style>
        @page {{ size: Letter; margin: 18mm 14mm; }}
        * {{ box-sizing: border-box; }}
        body {{ font-family: 'Helvetica', 'Arial', sans-serif; color: #000; font-size: 10.5pt; margin: 0; padding: 0; }}
        .header {{ text-align: center; border-bottom: 2px solid #000; padding-bottom: 8px; margin-bottom: 12px; }}
        .header h1 {{ margin: 0; font-size: 15pt; color: #000; }}
        .header .rnc {{ color: #444; font-size: 10pt; margin-top: 2px; }}
        .header .title {{ font-size: 12pt; font-weight: bold; margin-top: 4px; }}
        .header .date {{ color: #666; font-size: 9pt; margin-top: 2px; }}
        table {{ width: 100%; border-collapse: collapse; margin-top: 6px; }}
        thead th {{ background: #111; color: #fff; padding: 6px 8px; font-size: 10pt; text-align: left; }}
        thead th.money, thead th.num {{ text-align: right; }}
        td {{ padding: 4px 8px; font-size: 10pt; border-bottom: 1px solid #ddd; }}
        td.money {{ text-align: right; font-variant-numeric: tabular-nums; }}
        td.num {{ text-align: right; font-variant-numeric: tabular-nums; }}
        td.mono {{ font-family: 'Courier New', monospace; }}
        tr.row-peak td {{ border-top: 1px solid #000; border-bottom: 1px solid #000; font-weight: bold; }}
        tr.row-empty td {{ color: #777; }}
        tr.grand-total td {{ border-top: 2px solid #000; border-bottom: 2px solid #000; padding: 8px; background: #f5f5f5; font-size: 11pt; }}
        .footer {{ margin-top: 24px; text-align: center; color: #666; font-size: 8pt; font-style: italic; }}
        @media print {{ body {{ background: white; }} tr.grand-total td {{ background: #fff; }} }}
      </style>
    </head>
    <body>
      <div class="header">
        <h1>{business['name']}</h1>
        <div class="rnc">RNC: {business['rnc']}</div>
        <div class="title">Ventas por Hora</div>
        <div class="date">{period_label} · Generado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}</div>
      </div>
      <table>
        <thead>
          <tr>
            <th>Hora</th>
            <th class="num"># Facturas</th>
            <th class="money">Total Ventas</th>
            <th class="money">Ticket Promedio</th>
            <th class="num">% del Total</th>
          </tr>
        </thead>
        <tbody>
          {''.join(trs)}
        </tbody>
      </table>
      <div class="footer">
        Documento generado automáticamente por {business['name']} | {datetime.now().strftime('%d/%m/%Y %H:%M')} | Uso interno
      </div>
    </body>
    </html>
    """


@router.get("/hourly-sales/xlsx")
async def hourly_sales_xlsx(
    date_from: str = Query(...),
    date_to: str = Query(...),
    user=Depends(get_current_user),
):
    rows = await _fetch_hourly_data(date_from, date_to)
    business = await _get_business_info()
    buf = _build_hourly_xlsx(rows, _period_label(date_from, date_to), business)
    fname = f"VentasPorHora_{date_from}_al_{date_to}.xlsx"
    return _xlsx_response(buf, fname)


@router.get("/hourly-sales/pdf")
async def hourly_sales_pdf(
    date_from: str = Query(...),
    date_to: str = Query(...),
    user=Depends(get_current_user),
):
    try:
        from weasyprint import HTML
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"WeasyPrint no disponible: {e}")
    rows = await _fetch_hourly_data(date_from, date_to)
    business = await _get_business_info()
    html = _build_hourly_html(rows, _period_label(date_from, date_to), business)
    pdf_bytes = HTML(string=html).write_pdf()
    fname = f"VentasPorHora_{date_from}_al_{date_to}.pdf"
    return StreamingResponse(
        BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )



# ═══════════════════════════════════════════════════════════════
# IMPUESTOS — PDF + XLSX (breakdown ITBIS por tasa + Propina Legal 10%)
# ═══════════════════════════════════════════════════════════════

async def _fetch_taxes_data(date_from: str, date_to: str) -> dict:
    """Reuse the extended taxes_report function directly."""
    from routers.reports import taxes_report
    return await taxes_report(date_from=date_from, date_to=date_to)


def _build_taxes_xlsx(data: dict, period_label: str, business: dict) -> BytesIO:
    """XLSX with: Resumen por Tasa (SUM formulas), Propina Legal 10%, Desglose Diario (SUM)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Impuestos"

    BLACK_FILL = PatternFill("solid", fgColor="111111")
    WHITE_FONT = Font(bold=True, color="FFFFFF", size=11)
    BOLD = Font(bold=True, size=11)
    THIN_BLACK = Side(style="thin", color="000000")

    # Title block
    ws.cell(row=1, column=1, value=business.get("name", "VexlyPOS")).font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"RNC: {business.get('rnc','')}").font = Font(size=10, color="555555")
    ws.cell(row=3, column=1, value="Reporte de Impuestos").font = Font(bold=True, size=12)
    ws.cell(row=4, column=1, value=period_label).font = Font(size=10, color="555555")

    # ── Bloque 1: Resumen por Tasa (ITBIS 18%, 0%, Exento)
    ws.cell(row=6, column=1, value="RESUMEN POR TASA DE ITBIS").font = Font(bold=True, size=11)

    hdr_row = 7
    headers_tax = ["Tasa", "Base Imponible", "ITBIS Recaudado", "# Facturas"]
    for i, h in enumerate(headers_tax, start=1):
        c = ws.cell(row=hdr_row, column=i, value=h)
        c.font = WHITE_FONT
        c.fill = BLACK_FILL
        c.alignment = CENTER if i >= 2 else LEFT
        c.border = Border(top=THIN_BLACK, bottom=THIN_BLACK, left=THIN_BLACK, right=THIN_BLACK)

    br_start = hdr_row + 1
    breakdown = data.get("breakdown_by_rate", [])
    row_n = br_start
    for r in breakdown:
        ws.cell(row=row_n, column=1, value=r["rate_label"]).alignment = LEFT
        b = ws.cell(row=row_n, column=2, value=float(r["base"] or 0))
        b.number_format = '#,##0.00'; b.alignment = RIGHT
        i_c = ws.cell(row=row_n, column=3, value=float(r["itbis"] or 0))
        i_c.number_format = '#,##0.00'; i_c.alignment = RIGHT
        f_c = ws.cell(row=row_n, column=4, value=int(r["invoice_count"] or 0))
        f_c.alignment = RIGHT
        row_n += 1
    br_end = row_n - 1

    # TOTAL GENERAL row with SUM formulas
    total_tax_row = row_n
    t1 = ws.cell(row=total_tax_row, column=1, value="TOTAL GENERAL")
    t1.font = Font(bold=True, size=11)
    if breakdown:
        ws.cell(row=total_tax_row, column=2, value=f"=SUM(B{br_start}:B{br_end})")
        ws.cell(row=total_tax_row, column=3, value=f"=SUM(C{br_start}:C{br_end})")
        ws.cell(row=total_tax_row, column=4, value=f"=SUM(D{br_start}:D{br_end})")
    else:
        ws.cell(row=total_tax_row, column=2, value=0)
        ws.cell(row=total_tax_row, column=3, value=0)
        ws.cell(row=total_tax_row, column=4, value=0)
    for col, fmt in ((2, '#,##0.00'), (3, '#,##0.00'), (4, '#,##0')):
        cc = ws.cell(row=total_tax_row, column=col)
        cc.font = BOLD
        cc.number_format = fmt
        cc.alignment = RIGHT
        cc.border = Border(top=Side(style="medium", color="000000"), bottom=Side(style="medium", color="000000"))
    row_n += 2

    # ── Bloque 2: Propina Legal 10%
    ws.cell(row=row_n, column=1, value="PROPINA LEGAL 10%").font = Font(bold=True, size=11)
    row_n += 1
    summary = data.get("summary", {})
    total_subtotal = float(summary.get("total_subtotal") or 0)
    total_tips = float(summary.get("total_tips") or 0)

    ws.cell(row=row_n, column=1, value="Base gravable").alignment = LEFT
    bg = ws.cell(row=row_n, column=2, value=total_subtotal)
    bg.number_format = '#,##0.00'; bg.alignment = RIGHT
    row_n += 1

    ws.cell(row=row_n, column=1, value="Propina 10%").alignment = LEFT
    # Formula: =B{base_row}*0.10  (dynamic reference to base gravable row)
    pp = ws.cell(row=row_n, column=2, value=f"=B{row_n - 1}*0.10")
    pp.number_format = '#,##0.00'; pp.alignment = RIGHT; pp.font = BOLD
    # Also record actual value in a note (for audit) via comment-less cell C
    ws.cell(row=row_n, column=3, value=f"Valor calculado: {total_tips:,.2f}").font = Font(size=9, italic=True, color="555555")
    row_n += 2

    # ── Bloque 3: Desglose Diario
    ws.cell(row=row_n, column=1, value="DESGLOSE DIARIO").font = Font(bold=True, size=11)
    row_n += 1
    headers_daily = ["Fecha", "Subtotal", "ITBIS", "Propinas", "Total"]
    for i, h in enumerate(headers_daily, start=1):
        c = ws.cell(row=row_n, column=i, value=h)
        c.font = WHITE_FONT
        c.fill = BLACK_FILL
        c.alignment = CENTER if i >= 2 else LEFT
        c.border = Border(top=THIN_BLACK, bottom=THIN_BLACK, left=THIN_BLACK, right=THIN_BLACK)
    row_n += 1

    daily = data.get("daily", []) or []
    daily_start = row_n
    for d in daily:
        ws.cell(row=row_n, column=1, value=d.get("date", "")).font = Font(name="Courier New", size=10)
        for col, key in ((2, "subtotal"), (3, "itbis"), (4, "tips"), (5, "total")):
            c = ws.cell(row=row_n, column=col, value=float(d.get(key) or 0))
            c.number_format = '#,##0.00'; c.alignment = RIGHT
        row_n += 1
    daily_end = row_n - 1

    # Daily total row with SUM
    if daily:
        dt1 = ws.cell(row=row_n, column=1, value="TOTAL")
        dt1.font = Font(bold=True, size=11)
        for col, letter in ((2, "B"), (3, "C"), (4, "D"), (5, "E")):
            cc = ws.cell(row=row_n, column=col, value=f"=SUM({letter}{daily_start}:{letter}{daily_end})")
            cc.font = BOLD
            cc.number_format = '#,##0.00'
            cc.alignment = RIGHT
            cc.border = Border(top=Side(style="medium", color="000000"), bottom=Side(style="medium", color="000000"))
        row_n += 1

    # Integrity warning (if not ok)
    integrity = data.get("breakdown_integrity", {})
    if integrity and not integrity.get("ok"):
        row_n += 1
        ws.merge_cells(start_row=row_n, start_column=1, end_row=row_n, end_column=5)
        w = ws.cell(
            row=row_n, column=1,
            value=(
                f"⚠ Inconsistencia detectada: suma por tasa ({integrity.get('sum_by_rate')}) "
                f"≠ total ITBIS ({integrity.get('total_itbis')}). Diferencia: {integrity.get('diff')}."
            )
        )
        w.font = Font(bold=True, color="9C0006")
        w.fill = PatternFill("solid", fgColor="FFE0E0")

    # Footer
    row_n += 2
    ws.merge_cells(start_row=row_n, start_column=1, end_row=row_n, end_column=5)
    footer = ws.cell(
        row=row_n, column=1,
        value=f"Documento generado por {business.get('name','')} | {datetime.now().strftime('%d/%m/%Y %H:%M')} | Uso fiscal interno"
    )
    footer.font = Font(italic=True, size=9, color="555555")
    footer.alignment = CENTER

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 16
    ws.freeze_panes = "A8"

    buf = BytesIO()
    wb.save(buf); buf.seek(0)
    return buf


def _build_taxes_html(data: dict, period_label: str, business: dict) -> str:
    """Print-friendly B/W HTML for PDF generation."""
    breakdown = data.get("breakdown_by_rate", []) or []
    summary = data.get("summary", {}) or {}
    integrity = data.get("breakdown_integrity", {}) or {}

    total_base = sum(float(r.get("base") or 0) for r in breakdown)
    total_itbis_br = sum(float(r.get("itbis") or 0) for r in breakdown)
    total_invoices = sum(int(r.get("invoice_count") or 0) for r in breakdown)

    rows_tax = []
    for r in breakdown:
        rows_tax.append(
            f'<tr>'
            f'<td>{r["rate_label"]}</td>'
            f'<td class="money">{float(r["base"] or 0):,.2f}</td>'
            f'<td class="money">{float(r["itbis"] or 0):,.2f}</td>'
            f'<td class="num">{int(r["invoice_count"] or 0)}</td></tr>'
        )
    rows_tax.append(
        f'<tr class="grand-total">'
        f'<td><strong>TOTAL GENERAL</strong></td>'
        f'<td class="money"><strong>{total_base:,.2f}</strong></td>'
        f'<td class="money"><strong>{total_itbis_br:,.2f}</strong></td>'
        f'<td class="num"><strong>{total_invoices}</strong></td></tr>'
    )

    daily = data.get("daily", []) or []
    rows_daily = []
    td_sub = td_itb = td_tip = td_tot = 0.0
    for d in daily:
        td_sub += float(d.get("subtotal") or 0)
        td_itb += float(d.get("itbis") or 0)
        td_tip += float(d.get("tips") or 0)
        td_tot += float(d.get("total") or 0)
        rows_daily.append(
            f'<tr>'
            f'<td class="mono">{d.get("date","")}</td>'
            f'<td class="money">{float(d.get("subtotal") or 0):,.2f}</td>'
            f'<td class="money">{float(d.get("itbis") or 0):,.2f}</td>'
            f'<td class="money">{float(d.get("tips") or 0):,.2f}</td>'
            f'<td class="money">{float(d.get("total") or 0):,.2f}</td></tr>'
        )
    if daily:
        rows_daily.append(
            f'<tr class="grand-total">'
            f'<td><strong>TOTAL</strong></td>'
            f'<td class="money"><strong>{td_sub:,.2f}</strong></td>'
            f'<td class="money"><strong>{td_itb:,.2f}</strong></td>'
            f'<td class="money"><strong>{td_tip:,.2f}</strong></td>'
            f'<td class="money"><strong>{td_tot:,.2f}</strong></td></tr>'
        )

    base_gravable = float(summary.get("total_subtotal") or 0)
    propina_10 = round(base_gravable * 0.10, 2)

    integrity_banner = ""
    if integrity and not integrity.get("ok"):
        integrity_banner = (
            f'<div class="warning">⚠ Inconsistencia detectada: suma por tasa ({integrity.get("sum_by_rate")}) '
            f'≠ total ITBIS ({integrity.get("total_itbis")}). Diferencia: {integrity.get("diff")}.</div>'
        )

    return f"""
    <!DOCTYPE html>
    <html lang="es">
    <head>
      <meta charset="utf-8" />
      <title>Reporte de Impuestos</title>
      <style>
        @page {{ size: Letter; margin: 18mm 14mm; }}
        * {{ box-sizing: border-box; }}
        body {{ font-family: 'Helvetica', 'Arial', sans-serif; color: #000; font-size: 10.5pt; margin: 0; padding: 0; }}
        .header {{ text-align: center; border-bottom: 2px solid #000; padding-bottom: 8px; margin-bottom: 12px; }}
        .header h1 {{ margin: 0; font-size: 15pt; color: #000; }}
        .header .rnc {{ color: #444; font-size: 10pt; margin-top: 2px; }}
        .header .title {{ font-size: 12pt; font-weight: bold; margin-top: 4px; }}
        .header .date {{ color: #666; font-size: 9pt; margin-top: 2px; }}
        h2 {{ font-size: 11pt; margin: 14px 0 6px 0; border-bottom: 1px solid #000; padding-bottom: 3px; }}
        table {{ width: 100%; border-collapse: collapse; margin-top: 4px; }}
        thead th {{ background: #111; color: #fff; padding: 6px 8px; font-size: 10pt; text-align: left; }}
        thead th.money, thead th.num {{ text-align: right; }}
        td {{ padding: 4px 8px; font-size: 10pt; border-bottom: 1px solid #ddd; }}
        td.money, td.num {{ text-align: right; font-variant-numeric: tabular-nums; }}
        td.mono {{ font-family: 'Courier New', monospace; }}
        tr.grand-total td {{ border-top: 2px solid #000; border-bottom: 2px solid #000; padding: 7px 8px; background: #f5f5f5; font-size: 11pt; }}
        .tips-box {{ margin-top: 6px; border: 1px solid #000; padding: 8px 12px; display: inline-block; }}
        .tips-box p {{ margin: 3px 0; }}
        .tips-box strong {{ display: inline-block; min-width: 140px; }}
        .warning {{ background: #fff1f1; border: 1.5px solid #9C0006; color: #9C0006; padding: 8px 12px; margin-top: 10px; font-weight: bold; }}
        .footer {{ margin-top: 22px; text-align: center; color: #666; font-size: 8pt; font-style: italic; }}
        @media print {{ body {{ background: white; }} tr.grand-total td {{ background: #fff; }} }}
      </style>
    </head>
    <body>
      <div class="header">
        <h1>{business['name']}</h1>
        <div class="rnc">RNC: {business['rnc']}</div>
        <div class="title">Reporte de Impuestos</div>
        <div class="date">{period_label} · Generado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}</div>
      </div>

      {integrity_banner}

      <h2>Resumen por Tasa de ITBIS</h2>
      <table>
        <thead>
          <tr>
            <th>Tasa</th>
            <th class="money">Base Imponible</th>
            <th class="money">ITBIS Recaudado</th>
            <th class="num"># Facturas</th>
          </tr>
        </thead>
        <tbody>{''.join(rows_tax)}</tbody>
      </table>

      <h2>Propina Legal 10%</h2>
      <div class="tips-box">
        <p><strong>Base gravable:</strong> {base_gravable:,.2f}</p>
        <p><strong>Propina 10%:</strong> {propina_10:,.2f}</p>
      </div>

      <h2>Desglose Diario</h2>
      <table>
        <thead>
          <tr>
            <th>Fecha</th>
            <th class="money">Subtotal</th>
            <th class="money">ITBIS</th>
            <th class="money">Propinas</th>
            <th class="money">Total</th>
          </tr>
        </thead>
        <tbody>{''.join(rows_daily) or '<tr><td colspan="5" style="text-align:center;color:#777">Sin datos</td></tr>'}</tbody>
      </table>

      <div class="footer">
        Documento generado automáticamente por {business['name']} | {datetime.now().strftime('%d/%m/%Y %H:%M')} | Uso fiscal interno
      </div>
    </body>
    </html>
    """


@router.get("/taxes/xlsx")
async def taxes_xlsx(
    date_from: str = Query(...),
    date_to: str = Query(...),
    user=Depends(get_current_user),
):
    data = await _fetch_taxes_data(date_from, date_to)
    business = await _get_business_info()
    buf = _build_taxes_xlsx(data, _period_label(date_from, date_to), business)
    fname = f"ReporteImpuestos_{date_from}_al_{date_to}.xlsx"
    return _xlsx_response(buf, fname)


@router.get("/taxes/pdf")
async def taxes_pdf(
    date_from: str = Query(...),
    date_to: str = Query(...),
    user=Depends(get_current_user),
):
    try:
        from weasyprint import HTML
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"WeasyPrint no disponible: {e}")
    data = await _fetch_taxes_data(date_from, date_to)
    business = await _get_business_info()
    html = _build_taxes_html(data, _period_label(date_from, date_to), business)
    pdf_bytes = HTML(string=html).write_pdf()
    fname = f"ReporteImpuestos_{date_from}_al_{date_to}.pdf"
    return StreamingResponse(
        BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )



# ═══════════════════════════════════════════════════════════════
# CUENTAS ABIERTAS — PDF + XLSX (bills not yet paid)
# ═══════════════════════════════════════════════════════════════

async def _fetch_open_checks(date_from: str, date_to: str) -> dict:
    from routers.reports import _open_checks_impl
    return await _open_checks_impl(date_from=date_from, date_to=date_to)


def _fmt_short_dt(iso: str) -> str:
    if not iso:
        return "-"
    try:
        base = iso.split(".")[0].replace("Z", "")
        dt = datetime.fromisoformat(base)
        return dt.strftime("%d/%m/%Y %I:%M %p")
    except Exception:
        return iso


def _build_open_checks_xlsx(data: dict, period_label: str, business: dict) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Cuentas Abiertas"

    BLACK_FILL = PatternFill("solid", fgColor="111111")
    WHITE_FONT = Font(bold=True, color="FFFFFF", size=11)
    BOLD = Font(bold=True, size=11)
    THIN_BLACK = Side(style="thin", color="000000")

    # Title block
    ws.cell(row=1, column=1, value=business.get("name", "VexlyPOS")).font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"RNC: {business.get('rnc','')}").font = Font(size=10, color="555555")
    ws.cell(row=3, column=1, value="Cuentas Abiertas").font = Font(bold=True, size=12)
    ws.cell(row=4, column=1, value=period_label).font = Font(size=10, color="555555")

    # KPIs row (row 5)
    summary = data.get("summary", {}) or {}
    kpi_text = (
        f"Cuentas: {summary.get('count', 0)}  ·  "
        f"Monto en riesgo: {float(summary.get('total_value') or 0):,.2f}  ·  "
        f"Más antigua: {summary.get('oldest_minutes', 0)} min  ·  "
        f"Promedio: {summary.get('avg_minutes_open', 0)} min"
    )
    ws.cell(row=5, column=1, value=kpi_text).font = Font(size=10, italic=True, color="333333")

    # Header row 7
    headers = [
        "# Trans", "Mesa", "Mesero", "Abierta desde", "Min. abierta",
        "# Items", "Subtotal", "Total", "Estado",
    ]
    hdr_row = 7
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=hdr_row, column=i, value=h)
        c.font = WHITE_FONT
        c.fill = BLACK_FILL
        c.alignment = CENTER if i >= 5 else LEFT
        c.border = Border(top=THIN_BLACK, bottom=THIN_BLACK, left=THIN_BLACK, right=THIN_BLACK)

    bills = data.get("bills", []) or []
    row_n = hdr_row + 1
    data_start = row_n
    for b in bills:
        ws.cell(row=row_n, column=1, value=str(b.get("transaction_number") or "")).alignment = LEFT
        ws.cell(row=row_n, column=2, value=b.get("table", "")).alignment = LEFT
        ws.cell(row=row_n, column=3, value=b.get("waiter", "")).alignment = LEFT
        ws.cell(row=row_n, column=4, value=_fmt_short_dt(b.get("opened_at", ""))).alignment = LEFT
        ws.cell(row=row_n, column=5, value=int(b.get("minutes_open") or 0)).alignment = RIGHT
        ws.cell(row=row_n, column=6, value=int(b.get("items_count") or 0)).alignment = RIGHT
        c_sub = ws.cell(row=row_n, column=7, value=float(b.get("subtotal") or 0))
        c_sub.number_format = '#,##0.00'; c_sub.alignment = RIGHT
        c_tot = ws.cell(row=row_n, column=8, value=float(b.get("total") or 0))
        c_tot.number_format = '#,##0.00'; c_tot.alignment = RIGHT
        ws.cell(row=row_n, column=9, value=b.get("status", "")).alignment = CENTER
        row_n += 1
    data_end = row_n - 1

    # Totals row with SUM formulas
    total_row = row_n
    ws.cell(row=total_row, column=1, value=f"TOTAL [{len(bills)}]").font = BOLD
    if bills:
        ws.cell(row=total_row, column=5, value=f"=SUM(E{data_start}:E{data_end})")
        ws.cell(row=total_row, column=6, value=f"=SUM(F{data_start}:F{data_end})")
        ws.cell(row=total_row, column=7, value=f"=SUM(G{data_start}:G{data_end})")
        ws.cell(row=total_row, column=8, value=f"=SUM(H{data_start}:H{data_end})")
    else:
        for col in (5, 6, 7, 8):
            ws.cell(row=total_row, column=col, value=0)
    for col, fmt in ((5, '#,##0'), (6, '#,##0'), (7, '#,##0.00'), (8, '#,##0.00')):
        cc = ws.cell(row=total_row, column=col)
        cc.font = BOLD
        cc.number_format = fmt
        cc.alignment = RIGHT
        cc.border = Border(top=Side(style="medium", color="000000"), bottom=Side(style="medium", color="000000"))
    row_n += 2

    # ── By Waiter summary
    ws.cell(row=row_n, column=1, value="RESUMEN POR MESERO").font = Font(bold=True, size=11)
    row_n += 1
    for h_idx, h in enumerate(["Mesero", "# Cuentas", "Total"], start=1):
        c = ws.cell(row=row_n, column=h_idx, value=h)
        c.font = WHITE_FONT; c.fill = BLACK_FILL
        c.alignment = LEFT if h_idx == 1 else CENTER
    row_n += 1
    by_w = data.get("by_waiter", []) or []
    w_start = row_n
    for w in by_w:
        ws.cell(row=row_n, column=1, value=w.get("waiter", ""))
        ws.cell(row=row_n, column=2, value=int(w.get("count") or 0)).alignment = RIGHT
        ct = ws.cell(row=row_n, column=3, value=float(w.get("total") or 0))
        ct.number_format = '#,##0.00'; ct.alignment = RIGHT
        row_n += 1
    w_end = row_n - 1
    if by_w:
        ws.cell(row=row_n, column=1, value="TOTAL").font = BOLD
        ws.cell(row=row_n, column=2, value=f"=SUM(B{w_start}:B{w_end})")
        ws.cell(row=row_n, column=3, value=f"=SUM(C{w_start}:C{w_end})")
        for col, fmt in ((2, '#,##0'), (3, '#,##0.00')):
            cc = ws.cell(row=row_n, column=col)
            cc.font = BOLD; cc.number_format = fmt; cc.alignment = RIGHT
            cc.border = Border(top=Side(style="medium", color="000000"))

    # Footer
    row_n += 3
    ws.merge_cells(start_row=row_n, start_column=1, end_row=row_n, end_column=9)
    footer = ws.cell(
        row=row_n, column=1,
        value=f"Documento generado por {business.get('name','')} | {datetime.now().strftime('%d/%m/%Y %H:%M')} | Uso interno"
    )
    footer.font = Font(italic=True, size=9, color="555555")
    footer.alignment = CENTER

    # Widths
    widths = {"A": 14, "B": 18, "C": 22, "D": 22, "E": 12, "F": 10, "G": 14, "H": 14, "I": 12}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A8"

    buf = BytesIO()
    wb.save(buf); buf.seek(0)
    return buf


def _build_open_checks_html(data: dict, period_label: str, business: dict) -> str:
    summary = data.get("summary", {}) or {}
    bills = data.get("bills", []) or []
    by_waiter = data.get("by_waiter", []) or []

    trs = []
    total_minutes = total_items = 0
    total_subtotal = total_total = 0.0
    for b in bills:
        total_minutes += int(b.get("minutes_open") or 0)
        total_items += int(b.get("items_count") or 0)
        total_subtotal += float(b.get("subtotal") or 0)
        total_total += float(b.get("total") or 0)
        trs.append(
            f'<tr>'
            f'<td class="mono">{b.get("transaction_number","")}</td>'
            f'<td>{b.get("table","")}</td>'
            f'<td>{b.get("waiter","")}</td>'
            f'<td class="mono">{_fmt_short_dt(b.get("opened_at",""))}</td>'
            f'<td class="num">{int(b.get("minutes_open") or 0)}</td>'
            f'<td class="num">{int(b.get("items_count") or 0)}</td>'
            f'<td class="money">{float(b.get("subtotal") or 0):,.2f}</td>'
            f'<td class="money">{float(b.get("total") or 0):,.2f}</td>'
            f'<td class="center">{b.get("status","")}</td>'
            f'</tr>'
        )
    if bills:
        trs.append(
            f'<tr class="grand-total">'
            f'<td colspan="4"><strong>TOTAL [{len(bills)}]</strong></td>'
            f'<td class="num"><strong>{total_minutes}</strong></td>'
            f'<td class="num"><strong>{total_items}</strong></td>'
            f'<td class="money"><strong>{total_subtotal:,.2f}</strong></td>'
            f'<td class="money"><strong>{total_total:,.2f}</strong></td>'
            f'<td></td></tr>'
        )

    rows_waiter = []
    for w in by_waiter:
        rows_waiter.append(
            f'<tr>'
            f'<td>{w.get("waiter","")}</td>'
            f'<td class="num">{int(w.get("count") or 0)}</td>'
            f'<td class="money">{float(w.get("total") or 0):,.2f}</td>'
            f'</tr>'
        )

    return f"""
    <!DOCTYPE html>
    <html lang="es">
    <head>
      <meta charset="utf-8" />
      <title>Cuentas Abiertas</title>
      <style>
        @page {{ size: Letter landscape; margin: 14mm 12mm; }}
        * {{ box-sizing: border-box; }}
        body {{ font-family: 'Helvetica', 'Arial', sans-serif; color: #000; font-size: 10pt; margin: 0; padding: 0; }}
        .header {{ text-align: center; border-bottom: 2px solid #000; padding-bottom: 8px; margin-bottom: 10px; }}
        .header h1 {{ margin: 0; font-size: 15pt; color: #000; }}
        .header .rnc {{ color: #444; font-size: 9pt; margin-top: 2px; }}
        .header .title {{ font-size: 12pt; font-weight: bold; margin-top: 4px; }}
        .header .date {{ color: #666; font-size: 9pt; margin-top: 2px; }}
        .kpis {{ display: flex; gap: 20px; justify-content: center; margin: 8px 0 12px 0; font-size: 10pt; }}
        .kpis span {{ border: 1px solid #000; padding: 4px 10px; }}
        h2 {{ font-size: 11pt; margin: 12px 0 4px 0; border-bottom: 1px solid #000; padding-bottom: 3px; }}
        table {{ width: 100%; border-collapse: collapse; margin-top: 4px; }}
        thead th {{ background: #111; color: #fff; padding: 5px 6px; font-size: 9pt; text-align: left; }}
        thead th.money, thead th.num {{ text-align: right; }}
        td {{ padding: 3px 6px; font-size: 9pt; border-bottom: 1px solid #ddd; }}
        td.money, td.num {{ text-align: right; font-variant-numeric: tabular-nums; }}
        td.mono {{ font-family: 'Courier New', monospace; font-size: 9pt; }}
        td.center {{ text-align: center; }}
        tr.grand-total td {{ border-top: 2px solid #000; border-bottom: 2px solid #000; padding: 6px; background: #f5f5f5; font-size: 10pt; }}
        .footer {{ margin-top: 16px; text-align: center; color: #666; font-size: 8pt; font-style: italic; }}
        @media print {{ body {{ background: white; }} tr.grand-total td {{ background: #fff; }} }}
      </style>
    </head>
    <body>
      <div class="header">
        <h1>{business['name']}</h1>
        <div class="rnc">RNC: {business['rnc']}</div>
        <div class="title">Cuentas Abiertas</div>
        <div class="date">{period_label} · Generado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}</div>
      </div>

      <div class="kpis">
        <span><strong>Cuentas abiertas:</strong> {summary.get('count', 0)}</span>
        <span><strong>Monto en riesgo:</strong> {float(summary.get('total_value') or 0):,.2f}</span>
        <span><strong>Más antigua:</strong> {summary.get('oldest_minutes', 0)} min</span>
        <span><strong>Promedio:</strong> {summary.get('avg_minutes_open', 0)} min</span>
      </div>

      <h2>Detalle de Cuentas Abiertas</h2>
      <table>
        <thead>
          <tr>
            <th># Trans</th>
            <th>Mesa</th>
            <th>Mesero</th>
            <th>Abierta desde</th>
            <th class="num">Min.</th>
            <th class="num"># Items</th>
            <th class="money">Subtotal</th>
            <th class="money">Total</th>
            <th>Estado</th>
          </tr>
        </thead>
        <tbody>{''.join(trs) or '<tr><td colspan="9" style="text-align:center;color:#777">Sin cuentas abiertas en este período</td></tr>'}</tbody>
      </table>

      {('<h2>Resumen por Mesero</h2><table><thead><tr><th>Mesero</th><th class="num"># Cuentas</th><th class="money">Total</th></tr></thead><tbody>' + ''.join(rows_waiter) + '</tbody></table>') if rows_waiter else ''}

      <div class="footer">
        Documento generado automáticamente por {business['name']} | {datetime.now().strftime('%d/%m/%Y %H:%M')} | Uso interno
      </div>
    </body>
    </html>
    """


@router.get("/open-checks/xlsx")
async def open_checks_xlsx(
    date_from: str = Query(...),
    date_to: str = Query(...),
    user=Depends(get_current_user),
):
    data = await _fetch_open_checks(date_from, date_to)
    business = await _get_business_info()
    buf = _build_open_checks_xlsx(data, _period_label(date_from, date_to), business)
    fname = f"CuentasAbiertas_{date_from}_al_{date_to}.xlsx"
    return _xlsx_response(buf, fname)


@router.get("/open-checks/pdf")
async def open_checks_pdf(
    date_from: str = Query(...),
    date_to: str = Query(...),
    user=Depends(get_current_user),
):
    try:
        from weasyprint import HTML
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"WeasyPrint no disponible: {e}")
    data = await _fetch_open_checks(date_from, date_to)
    business = await _get_business_info()
    html = _build_open_checks_html(data, _period_label(date_from, date_to), business)
    pdf_bytes = HTML(string=html).write_pdf()
    fname = f"CuentasAbiertas_{date_from}_al_{date_to}.pdf"
    return StreamingResponse(
        BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )



# ═══════════════════════════════════════════════════════════════
# VENTAS COMPARATIVAS — PDF + XLSX (Periodo A vs Periodo B)
# ═══════════════════════════════════════════════════════════════

async def _fetch_sales_comparative(pa_from, pa_to, pb_from, pb_to) -> dict:
    from routers.reports import _sales_comparative_impl
    return await _sales_comparative_impl(pa_from, pa_to, pb_from, pb_to)


def _build_sales_comparative_xlsx(data: dict, business: dict) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparativo"

    BLACK_FILL = PatternFill("solid", fgColor="111111")
    WHITE_FONT = Font(bold=True, color="FFFFFF", size=11)
    BOLD = Font(bold=True, size=11)
    THIN_BLACK = Side(style="thin", color="000000")

    pa = data.get("period_a", {}) or {}
    pb = data.get("period_b", {}) or {}
    metrics = data.get("metrics", []) or []

    # Title block
    ws.cell(row=1, column=1, value=business.get("name", "VexlyPOS")).font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"RNC: {business.get('rnc','')}").font = Font(size=10, color="555555")
    ws.cell(row=3, column=1, value="Ventas Comparativas — Período A vs Período B").font = Font(bold=True, size=12)
    period_a_label = f"A: {pa.get('date_from','')} al {pa.get('date_to','')}"
    period_b_label = f"B: {pb.get('date_from','')} al {pb.get('date_to','')}"
    ws.cell(row=4, column=1, value=f"{period_a_label}   ·   {period_b_label}").font = Font(size=10, color="555555")

    # Header row 6
    headers = ["Métrica", period_a_label, period_b_label, "Diferencia", "% Cambio"]
    hdr_row = 6
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=hdr_row, column=i, value=h)
        c.font = WHITE_FONT
        c.fill = BLACK_FILL
        c.alignment = CENTER if i >= 2 else LEFT
        c.border = Border(top=THIN_BLACK, bottom=THIN_BLACK, left=THIN_BLACK, right=THIN_BLACK)

    row_n = hdr_row + 1
    for m in metrics:
        kind = m.get("kind", "money")
        ws.cell(row=row_n, column=1, value=m["label"]).alignment = LEFT
        b_cell = ws.cell(row=row_n, column=2, value=float(m.get("period_a") or 0))
        c_cell = ws.cell(row=row_n, column=3, value=float(m.get("period_b") or 0))
        # Diff formula: =C-B (real formula, not static value)
        d_cell = ws.cell(row=row_n, column=4, value=f"=C{row_n}-B{row_n}")
        # % formula: =IF(B=0,"N/A",(C-B)/B)
        e_cell = ws.cell(row=row_n, column=5, value=f'=IF(B{row_n}=0,"N/A",(C{row_n}-B{row_n})/B{row_n})')

        fmt_num = '#,##0' if kind == "int" else '#,##0.00'
        for cc in (b_cell, c_cell, d_cell):
            cc.number_format = fmt_num
            cc.alignment = RIGHT
        e_cell.number_format = '0.00%'
        e_cell.alignment = RIGHT
        # Highlight positive/negative direction via bold on diff
        d_cell.font = Font(bold=True, color="1F7A1F" if float(m.get("diff") or 0) >= 0 else "9C0006")
        row_n += 1

    # Footer
    row_n += 2
    ws.merge_cells(start_row=row_n, start_column=1, end_row=row_n, end_column=5)
    footer = ws.cell(
        row=row_n, column=1,
        value=f"Documento generado por {business.get('name','')} | {datetime.now().strftime('%d/%m/%Y %H:%M')} | Uso interno"
    )
    footer.font = Font(italic=True, size=9, color="555555")
    footer.alignment = CENTER

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 12
    ws.freeze_panes = "A7"

    buf = BytesIO()
    wb.save(buf); buf.seek(0)
    return buf


def _build_sales_comparative_html(data: dict, business: dict) -> str:
    pa = data.get("period_a", {}) or {}
    pb = data.get("period_b", {}) or {}
    metrics = data.get("metrics", []) or []

    def _fmt(v, kind):
        if kind == "int":
            return f"{int(v):,}"
        return f"{float(v):,.2f}"

    def _fmt_pct(pct):
        if pct is None:
            return "N/A"
        sign = "▲" if pct >= 0 else "▼"
        return f"{sign} {pct:+.2f}%"

    trs = []
    for m in metrics:
        pct = m.get("pct")
        diff = m.get("diff", 0)
        color_cls = "pos" if diff >= 0 else "neg"
        trs.append(
            f'<tr>'
            f'<td>{m["label"]}</td>'
            f'<td class="money">{_fmt(m["period_a"], m["kind"])}</td>'
            f'<td class="money">{_fmt(m["period_b"], m["kind"])}</td>'
            f'<td class="money {color_cls}"><strong>{_fmt(diff, m["kind"])}</strong></td>'
            f'<td class="num {color_cls}"><strong>{_fmt_pct(pct)}</strong></td>'
            f'</tr>'
        )

    return f"""
    <!DOCTYPE html>
    <html lang="es">
    <head>
      <meta charset="utf-8" />
      <title>Ventas Comparativas</title>
      <style>
        @page {{ size: Letter; margin: 18mm 14mm; }}
        * {{ box-sizing: border-box; }}
        body {{ font-family: 'Helvetica', 'Arial', sans-serif; color: #000; font-size: 10.5pt; margin: 0; padding: 0; }}
        .header {{ text-align: center; border-bottom: 2px solid #000; padding-bottom: 8px; margin-bottom: 12px; }}
        .header h1 {{ margin: 0; font-size: 15pt; color: #000; }}
        .header .rnc {{ color: #444; font-size: 10pt; margin-top: 2px; }}
        .header .title {{ font-size: 12pt; font-weight: bold; margin-top: 4px; }}
        .header .date {{ color: #666; font-size: 9pt; margin-top: 2px; }}
        .periods {{ display: flex; gap: 14px; margin: 10px 0; }}
        .periods .box {{ flex: 1; border: 1.5px solid #000; padding: 8px 10px; }}
        .periods .box .hd {{ font-weight: bold; font-size: 10pt; margin-bottom: 3px; }}
        .periods .box .rg {{ color: #333; font-size: 9pt; font-family: 'Courier New', monospace; }}
        table {{ width: 100%; border-collapse: collapse; margin-top: 10px; }}
        thead th {{ background: #111; color: #fff; padding: 6px 8px; font-size: 10pt; text-align: left; }}
        thead th.money, thead th.num {{ text-align: right; }}
        td {{ padding: 5px 8px; font-size: 10pt; border-bottom: 1px solid #ddd; }}
        td.money, td.num {{ text-align: right; font-variant-numeric: tabular-nums; }}
        td.pos {{ /* kept monochrome in print; visual hint preserved via bold */ }}
        td.neg {{ /* kept monochrome in print */ }}
        .footer {{ margin-top: 24px; text-align: center; color: #666; font-size: 8pt; font-style: italic; }}
        @media print {{ body {{ background: white; }} }}
      </style>
    </head>
    <body>
      <div class="header">
        <h1>{business['name']}</h1>
        <div class="rnc">RNC: {business['rnc']}</div>
        <div class="title">Ventas Comparativas — Período A vs Período B</div>
        <div class="date">Generado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}</div>
      </div>

      <div class="periods">
        <div class="box">
          <div class="hd">Período A</div>
          <div class="rg">{pa.get('date_from','')} al {pa.get('date_to','')}</div>
        </div>
        <div class="box">
          <div class="hd">Período B</div>
          <div class="rg">{pb.get('date_from','')} al {pb.get('date_to','')}</div>
        </div>
      </div>

      <table>
        <thead>
          <tr>
            <th>Métrica</th>
            <th class="money">Período A</th>
            <th class="money">Período B</th>
            <th class="money">Diferencia</th>
            <th class="num">% Cambio</th>
          </tr>
        </thead>
        <tbody>{''.join(trs) or '<tr><td colspan="5" style="text-align:center;color:#777">Sin datos</td></tr>'}</tbody>
      </table>

      <div class="footer">
        Documento generado automáticamente por {business['name']} | {datetime.now().strftime('%d/%m/%Y %H:%M')} | Uso interno
      </div>
    </body>
    </html>
    """


@router.get("/sales-comparative/xlsx")
async def sales_comparative_xlsx(
    period_a_from: str = Query(...),
    period_a_to: str = Query(...),
    period_b_from: str = Query(...),
    period_b_to: str = Query(...),
    user=Depends(get_current_user),
):
    data = await _fetch_sales_comparative(period_a_from, period_a_to, period_b_from, period_b_to)
    business = await _get_business_info()
    buf = _build_sales_comparative_xlsx(data, business)
    fname = f"VentasComparativas_A{period_a_from}_B{period_b_from}.xlsx"
    return _xlsx_response(buf, fname)


@router.get("/sales-comparative/pdf")
async def sales_comparative_pdf(
    period_a_from: str = Query(...),
    period_a_to: str = Query(...),
    period_b_from: str = Query(...),
    period_b_to: str = Query(...),
    user=Depends(get_current_user),
):
    try:
        from weasyprint import HTML
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"WeasyPrint no disponible: {e}")
    data = await _fetch_sales_comparative(period_a_from, period_a_to, period_b_from, period_b_to)
    business = await _get_business_info()
    html = _build_sales_comparative_html(data, business)
    pdf_bytes = HTML(string=html).write_pdf()
    fname = f"VentasComparativas_A{period_a_from}_B{period_b_from}.pdf"
    return StreamingResponse(
        BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )

