"""
Pytest suite for extended Taxes Report (ITBIS breakdown by rate for DGII IT-1)
Covers:
- /api/reports/taxes returns new fields breakdown_by_rate + breakdown_integrity, keeps legacy fields.
- Integrity: sum(by_rate) == total_itbis for seeded data 2026-04-01..2026-04-20 (expected ITBIS 4518).
- /api/reports/xlsx/taxes/xlsx → XLSX with formulas (=SUM, =B{row}*0.10, =SUM on daily TOTAL row).
- /api/reports/xlsx/taxes/pdf → valid PDF with Resumen por Tasa + Propina Legal + Desglose Diario text.
- Regression: /api/reports/taxes without params still returns full structure.
- Regression: endpoints for Prompts 1/2/3 keep responding OK.
- Auth enforcement on export endpoints.
"""
import os
import io
import pytest
import requests
from openpyxl import load_workbook

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "https://factura-consumo-app.preview.emergentagent.com").rstrip("/")
ADMIN_PIN = "11338585"

RANGE_FROM = "2026-04-01"
RANGE_TO = "2026-04-20"


@pytest.fixture(scope="module")
def token():
    r = requests.post(f"{BASE_URL}/api/auth/login", json={"pin": ADMIN_PIN}, timeout=20)
    assert r.status_code == 200, r.text
    tok = r.json().get("token")
    assert tok
    return tok


@pytest.fixture(scope="module")
def auth_headers(token):
    return {"Authorization": f"Bearer {token}"}


# ── JSON endpoint ─────────────────────────────────────────────
class TestTaxesJSON:
    def test_taxes_with_range_has_breakdown(self, auth_headers):
        r = requests.get(
            f"{BASE_URL}/api/reports/taxes",
            params={"date_from": RANGE_FROM, "date_to": RANGE_TO},
            headers=auth_headers,
            timeout=30,
        )
        assert r.status_code == 200, r.text
        data = r.json()
        # Legacy fields preserved
        assert "summary" in data and "daily" in data
        s = data["summary"]
        for k in ("total_itbis", "total_tips", "total_sales", "total_subtotal"):
            assert k in s, f"missing legacy field {k}"
        # New fields
        assert "breakdown_by_rate" in data
        assert "breakdown_integrity" in data
        assert isinstance(data["breakdown_by_rate"], list)
        integ = data["breakdown_integrity"]
        for k in ("ok", "sum_by_rate", "total_itbis", "diff"):
            assert k in integ

    def test_taxes_integrity_ok_and_values(self, auth_headers):
        r = requests.get(
            f"{BASE_URL}/api/reports/taxes",
            params={"date_from": RANGE_FROM, "date_to": RANGE_TO},
            headers=auth_headers,
            timeout=30,
        )
        data = r.json()
        integ = data["breakdown_integrity"]
        assert integ["ok"] is True, f"integrity not ok: {integ}"
        assert abs(integ["diff"]) <= 0.02
        # Expected from problem statement: total ITBIS 4518
        assert round(data["summary"]["total_itbis"], 2) == 4518.0
        assert round(integ["sum_by_rate"], 2) == 4518.0
        # breakdown 18% with base 25100, itbis 4518, 42 invoices
        b18 = [x for x in data["breakdown_by_rate"] if "18" in x["rate_label"]]
        assert len(b18) == 1, f"expected one 18% bucket, got {data['breakdown_by_rate']}"
        row = b18[0]
        assert round(row["base"], 2) == 25100.0
        assert round(row["itbis"], 2) == 4518.0
        assert row["invoice_count"] == 42

    def test_taxes_no_params_regression(self, auth_headers):
        r = requests.get(f"{BASE_URL}/api/reports/taxes", headers=auth_headers, timeout=30)
        assert r.status_code == 200
        data = r.json()
        for k in ("summary", "daily", "breakdown_by_rate", "breakdown_integrity"):
            assert k in data


# ── XLSX export ──────────────────────────────────────────────
class TestTaxesXLSX:
    def test_xlsx_download_valid(self, auth_headers):
        r = requests.get(
            f"{BASE_URL}/api/reports/xlsx/taxes/xlsx",
            params={"date_from": RANGE_FROM, "date_to": RANGE_TO},
            headers=auth_headers,
            timeout=60,
        )
        assert r.status_code == 200, r.text[:400]
        assert r.content[:2] == b"PK", "not a valid XLSX (missing PK magic)"
        wb = load_workbook(io.BytesIO(r.content), data_only=False)
        ws = wb.active
        # Scan cells for formulas we expect
        found_sum_formula = False
        found_tip_formula = False
        found_daily_sum_formula = False
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and v.startswith("="):
                    vs = v.upper()
                    if "SUM(" in vs:
                        found_sum_formula = True
                        # Daily TOTAL row references spans of cells across columns B..E
                        if any(col in vs for col in ("B", "C", "D", "E")) and ":" in vs:
                            found_daily_sum_formula = True
                    if "*0.1" in v or "*0.10" in v:
                        found_tip_formula = True
        assert found_sum_formula, "No =SUM(...) formula found in XLSX"
        assert found_tip_formula, "No =B{row}*0.10 tip formula found in XLSX (must not be hardcoded)"
        assert found_daily_sum_formula, "Expected =SUM(...) on daily TOTAL row across columns"


# ── PDF export ───────────────────────────────────────────────
class TestTaxesPDF:
    def test_pdf_download_valid(self, auth_headers):
        r = requests.get(
            f"{BASE_URL}/api/reports/xlsx/taxes/pdf",
            params={"date_from": RANGE_FROM, "date_to": RANGE_TO},
            headers=auth_headers,
            timeout=90,
        )
        assert r.status_code == 200, r.text[:400]
        assert r.content[:4] == b"%PDF", "not a valid PDF"
        # Extract text to confirm sections
        try:
            from pypdf import PdfReader
        except ImportError:
            pytest.skip("pypdf not available")
        reader = PdfReader(io.BytesIO(r.content))
        text = "".join((p.extract_text() or "") for p in reader.pages)
        assert "Resumen por Tasa" in text or "Tasa" in text, f"PDF missing Resumen por Tasa: {text[:400]}"
        assert "Propina" in text, "PDF missing Propina Legal"
        assert "Desglose Diario" in text or "Diario" in text, "PDF missing Desglose Diario"


# ── Auth enforcement ────────────────────────────────────────
class TestAuth:
    def test_xlsx_requires_auth(self):
        r = requests.get(
            f"{BASE_URL}/api/reports/xlsx/taxes/xlsx",
            params={"date_from": RANGE_FROM, "date_to": RANGE_TO},
            timeout=30,
        )
        assert r.status_code in (401, 403), f"expected auth error, got {r.status_code}"

    def test_pdf_requires_auth(self):
        r = requests.get(
            f"{BASE_URL}/api/reports/xlsx/taxes/pdf",
            params={"date_from": RANGE_FROM, "date_to": RANGE_TO},
            timeout=30,
        )
        assert r.status_code in (401, 403)


# ── Regression on other reports ─────────────────────────────
class TestRegression:
    def test_category_sales(self, auth_headers):
        r = requests.get(
            f"{BASE_URL}/api/reports/sales-by-category",
            params={"date_from": RANGE_FROM, "date_to": RANGE_TO},
            headers=auth_headers,
            timeout=30,
        )
        assert r.status_code == 200

    def test_cierre_caja(self, auth_headers):
        r = requests.get(
            f"{BASE_URL}/api/reports/cash-close-hierarchical",
            params={"date_from": RANGE_FROM, "date_to": RANGE_TO},
            headers=auth_headers,
            timeout=30,
        )
        assert r.status_code == 200

    def test_hourly_sales(self, auth_headers):
        r = requests.get(
            f"{BASE_URL}/api/reports/hourly-sales",
            params={"date_from": RANGE_FROM, "date_to": RANGE_TO},
            headers=auth_headers,
            timeout=30,
        )
        assert r.status_code == 200

    def test_hourly_pdf(self, auth_headers):
        r = requests.get(
            f"{BASE_URL}/api/reports/xlsx/hourly-sales/pdf",
            params={"date_from": RANGE_FROM, "date_to": RANGE_TO},
            headers=auth_headers,
            timeout=60,
        )
        assert r.status_code == 200 and r.content[:4] == b"%PDF"

    def test_hourly_xlsx(self, auth_headers):
        r = requests.get(
            f"{BASE_URL}/api/reports/xlsx/hourly-sales/xlsx",
            params={"date_from": RANGE_FROM, "date_to": RANGE_TO},
            headers=auth_headers,
            timeout=60,
        )
        assert r.status_code == 200 and r.content[:2] == b"PK"
