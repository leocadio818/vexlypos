"""
Tests for A5 Sales by Table / Area report.
Endpoint: GET /api/reports/sales-by-table
XLSX:     GET /api/reports/xlsx/sales-by-table/xlsx
PDF:      GET /api/reports/xlsx/sales-by-table/pdf
"""
import os
import io
import pytest
import requests
from openpyxl import load_workbook

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL").rstrip("/")
ADMIN_PIN = "11338585"
DATE_FROM = "2026-04-01"
DATE_TO = "2026-04-20"


@pytest.fixture(scope="module")
def token():
    r = requests.post(
        f"{BASE_URL}/api/auth/login",
        json={"pin": ADMIN_PIN},
        timeout=20,
    )
    assert r.status_code == 200, f"login failed: {r.status_code} {r.text[:200]}"
    data = r.json()
    t = data.get("token") or data.get("access_token")
    assert t, f"no token in response: {data}"
    return t


@pytest.fixture(scope="module")
def auth_headers(token):
    return {"Authorization": f"Bearer {token}"}


# ---------- JSON endpoint ----------


class TestSalesByTableJSON:
    def test_json_structure_and_values(self, auth_headers):
        r = requests.get(
            f"{BASE_URL}/api/reports/sales-by-table",
            params={"date_from": DATE_FROM, "date_to": DATE_TO, "limit": 20},
            headers=auth_headers,
            timeout=30,
        )
        assert r.status_code == 200, r.text[:300]
        data = r.json()
        assert "summary" in data and "top_tables" in data and "by_area" in data

        s = data["summary"]
        # Expected totals per review request
        assert s.get("grand_total") == 32144, f"grand_total={s.get('grand_total')}"
        assert s.get("grand_bills") == 42, f"grand_bills={s.get('grand_bills')}"
        assert s.get("tables_with_sales") == 12
        assert s.get("areas_with_sales") == 3

        # Top table
        top_tables = data["top_tables"]
        assert len(top_tables) >= 1
        first = top_tables[0]
        assert first["table"] == "Mesa 1"
        assert first["area"] == "Terraza"
        assert first["total"] == 8208
        assert first["bills"] == 16
        assert abs(first["pct"] - 25.54) < 0.1, f"pct={first['pct']}"

        # By area: Salon Principal top
        by_area = data["by_area"]
        assert len(by_area) >= 3
        area_map = {a["area"]: a for a in by_area}
        sp = area_map.get("Salon Principal")
        assert sp is not None
        assert sp["total"] == 17792
        assert sp["bills"] == 22
        assert sp["tables_count"] == 7
        assert sp["top_table"] == "Mesa 4"

        terraza = area_map.get("Terraza")
        assert terraza and terraza["total"] == 8976

        vip = area_map.get("VIP")
        assert vip and vip["total"] == 5376
        assert vip["top_table"] == "Mesa 21"

    def test_limit_parameter(self, auth_headers):
        r = requests.get(
            f"{BASE_URL}/api/reports/sales-by-table",
            params={"date_from": DATE_FROM, "date_to": DATE_TO, "limit": 5},
            headers=auth_headers,
            timeout=30,
        )
        assert r.status_code == 200
        assert len(r.json()["top_tables"]) <= 5


# ---------- Auth enforcement on XLSX/PDF ----------


class TestSalesByTableAuth:
    def test_xlsx_without_auth_blocked(self):
        r = requests.get(
            f"{BASE_URL}/api/reports/xlsx/sales-by-table/xlsx",
            params={"date_from": DATE_FROM, "date_to": DATE_TO},
            timeout=30,
        )
        assert r.status_code in (401, 403), f"got {r.status_code}"

    def test_pdf_without_auth_blocked(self):
        r = requests.get(
            f"{BASE_URL}/api/reports/xlsx/sales-by-table/pdf",
            params={"date_from": DATE_FROM, "date_to": DATE_TO},
            timeout=30,
        )
        assert r.status_code in (401, 403)


# ---------- PDF export ----------


class TestSalesByTablePDF:
    def test_pdf_returns_valid_pdf(self, auth_headers):
        r = requests.get(
            f"{BASE_URL}/api/reports/xlsx/sales-by-table/pdf",
            params={"date_from": DATE_FROM, "date_to": DATE_TO},
            headers=auth_headers,
            timeout=60,
        )
        assert r.status_code == 200, r.text[:300]
        assert r.content[:4] == b"%PDF", f"magic={r.content[:8]!r}"
        assert len(r.content) > 1000


# ---------- XLSX export with formula introspection ----------


class TestSalesByTableXLSX:
    @pytest.fixture(scope="class")
    def wb(self, auth_headers):
        r = requests.get(
            f"{BASE_URL}/api/reports/xlsx/sales-by-table/xlsx",
            params={"date_from": DATE_FROM, "date_to": DATE_TO},
            headers=auth_headers,
            timeout=60,
        )
        assert r.status_code == 200, r.text[:300]
        assert r.content[:2] == b"PK", f"magic={r.content[:4]!r}"
        assert len(r.content) > 1000
        return load_workbook(io.BytesIO(r.content))

    def test_xlsx_formulas_present(self, wb):
        ws = wb.active
        # Dump all cells to find formulas
        rows = {}
        for row in ws.iter_rows(values_only=False):
            for c in row:
                if c.value is not None:
                    rows[c.coordinate] = c.value

        # Collect formulas (strings starting with '=')
        formulas = {k: v for k, v in rows.items() if isinstance(v, str) and v.startswith("=")}
        assert formulas, "No formulas found in XLSX — must be real formulas not hardcoded"

        fstrs = " | ".join(formulas.values())
        # TOP MESAS section must have =SUM for bills, guests, total (cols D,E,F per spec)
        assert "=SUM(D" in fstrs, f"missing =SUM(D..D): {fstrs[:400]}"
        assert "=SUM(E" in fstrs, f"missing =SUM(E..E): {fstrs[:400]}"
        assert "=SUM(F" in fstrs, f"missing =SUM(F..F): {fstrs[:400]}"

        # ticket prom in TOTAL TOP row should be =IF(D{row}>0,F{row}/D{row},0)
        has_ticket_prom_if = any(
            (v.startswith("=IF(") and ("/D" in v or "/E" in v) and ">0" in v)
            for v in formulas.values()
        )
        assert has_ticket_prom_if, f"missing ticket prom IF formula. formulas={list(formulas.values())[:20]}"

        # % per row should reference F column (e.g., =IF(F{total}>0,F{row}/F{total},0))
        has_pct = any(
            (v.startswith("=IF(") and "F" in v and ">0" in v and "/F" in v)
            for v in formulas.values()
        )
        assert has_pct, f"missing % formulas: {list(formulas.values())[:20]}"

    def test_xlsx_by_area_sum_formulas(self, wb):
        ws = wb.active
        formulas = []
        for row in ws.iter_rows(values_only=False):
            for c in row:
                if isinstance(c.value, str) and c.value.startswith("="):
                    formulas.append(c.value)

        # RANKING POR ÁREA row TOTAL should contain =SUM(B..B), =SUM(C..C), =SUM(D..D)
        fstrs = " ".join(formulas)
        assert "=SUM(B" in fstrs
        assert "=SUM(C" in fstrs
        assert "=SUM(D" in fstrs
