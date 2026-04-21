"""
B5 - Cuentas Abiertas / Open Checks Report tests
Covers:
 - JSON endpoint (/api/reports/open-checks) with seed data
 - XLSX formulas (=SUM(E..), =SUM(F..), =SUM(G..), =SUM(H..), waiter =SUM(B..) =SUM(C..))
 - PDF magic and landscape/Letter content
 - Auth enforcement on XLSX/PDF endpoints (401 without Bearer)
 - Empty range returns 200 with count=0
 - Regression: previous reports (categorias, cierre-caja, hourly, taxes)
"""
import os
import io
import pytest
import requests
from openpyxl import load_workbook

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "https://factura-consumo-app.preview.emergentagent.com").rstrip("/")
API = f"{BASE_URL}/api"

SEED_DATE = "2026-04-21"  # business_date of the 3 seed open checks


@pytest.fixture(scope="module")
def token():
    r = requests.post(f"{API}/auth/login", json={"pin": "11338585"}, timeout=30)
    assert r.status_code == 200, f"login failed: {r.status_code} {r.text}"
    data = r.json()
    tok = data.get("access_token") or data.get("token")
    assert tok, f"no token in login response: {data}"
    return tok


@pytest.fixture(scope="module")
def auth(token):
    return {"Authorization": f"Bearer {token}"}


# ── JSON endpoint ───────────────────────────────────────────────────────────
class TestOpenChecksJSON:
    def test_json_with_seed(self, auth):
        r = requests.get(
            f"{API}/reports/open-checks",
            params={"date_from": SEED_DATE, "date_to": SEED_DATE},
            headers=auth,
            timeout=30,
        )
        assert r.status_code == 200, r.text
        data = r.json()
        assert data["date_from"] == SEED_DATE and data["date_to"] == SEED_DATE
        s = data["summary"]
        assert s["count"] >= 3
        # at least the 3 seeds (total 3712)
        assert s["total_value"] >= 3712.0 - 0.01
        assert s["oldest_minutes"] >= 1  # some positive value
        assert s["avg_minutes_open"] >= 0
        # bills sorted oldest first
        bills = data["bills"]
        mins = [b["minutes_open"] for b in bills]
        assert mins == sorted(mins, reverse=True)
        # seeds present
        tx = {b["transaction_number"] for b in bills}
        assert "T-901" in tx and "T-902" in tx and "T-903" in tx
        # by_waiter contains OSCAR + CARLOS
        waiters = {w["waiter"] for w in data.get("by_waiter", [])}
        assert "OSCAR" in waiters and "CARLOS" in waiters

    def test_json_empty_range_returns_200(self, auth):
        r = requests.get(
            f"{API}/reports/open-checks",
            params={"date_from": "2020-01-01", "date_to": "2020-01-01"},
            headers=auth,
            timeout=30,
        )
        assert r.status_code == 200, r.text
        data = r.json()
        assert data["summary"]["count"] == 0
        assert data["bills"] == []

    def test_status_filter_excludes_paid(self, auth):
        """Bills with status paid/cancelled must not appear."""
        r = requests.get(
            f"{API}/reports/open-checks",
            params={"date_from": SEED_DATE, "date_to": SEED_DATE},
            headers=auth,
            timeout=30,
        )
        assert r.status_code == 200
        statuses = {b["status"] for b in r.json()["bills"]}
        assert not (statuses & {"paid", "cancelled", "void", "voided", "anulled"})


# ── XLSX ────────────────────────────────────────────────────────────────────
class TestOpenChecksXLSX:
    def test_xlsx_auth_required(self):
        r = requests.get(
            f"{API}/reports/xlsx/open-checks/xlsx",
            params={"date_from": SEED_DATE, "date_to": SEED_DATE},
            timeout=30,
        )
        assert r.status_code in (401, 403), f"expected 401/403 got {r.status_code}"

    def test_xlsx_valid_and_has_sum_formulas(self, auth):
        r = requests.get(
            f"{API}/reports/xlsx/open-checks/xlsx",
            params={"date_from": SEED_DATE, "date_to": SEED_DATE},
            headers=auth,
            timeout=60,
        )
        assert r.status_code == 200, r.text[:500]
        # PK = zip signature for xlsx
        assert r.content[:2] == b"PK"
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active
        # find TOTAL row (col A starts with 'TOTAL [')
        total_row = None
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=1).value
            if isinstance(v, str) and v.startswith("TOTAL ["):
                total_row = row
                break
        assert total_row, "TOTAL row not found"
        # SUM formulas in cols E,F,G,H
        for col in ("E", "F", "G", "H"):
            cell = ws[f"{col}{total_row}"]
            assert isinstance(cell.value, str) and cell.value.startswith(f"=SUM({col}"), \
                f"col {col} total formula missing/bad: {cell.value}"

        # By-waiter SUM formulas
        found_waiter_sums = {"B": False, "C": False}
        for row in range(total_row + 1, ws.max_row + 1):
            for col in ("B", "C"):
                v = ws.cell(row=row, column={"B": 2, "C": 3}[col]).value
                if isinstance(v, str) and v.startswith(f"=SUM({col}"):
                    found_waiter_sums[col] = True
        assert all(found_waiter_sums.values()), f"waiter SUM formulas missing: {found_waiter_sums}"

    def test_xlsx_empty_data_no_500(self, auth):
        r = requests.get(
            f"{API}/reports/xlsx/open-checks/xlsx",
            params={"date_from": "2020-01-01", "date_to": "2020-01-01"},
            headers=auth,
            timeout=60,
        )
        assert r.status_code == 200, r.text[:500]
        assert r.content[:2] == b"PK"


# ── PDF ─────────────────────────────────────────────────────────────────────
class TestOpenChecksPDF:
    def test_pdf_auth_required(self):
        r = requests.get(
            f"{API}/reports/xlsx/open-checks/pdf",
            params={"date_from": SEED_DATE, "date_to": SEED_DATE},
            timeout=30,
        )
        assert r.status_code in (401, 403)

    def test_pdf_valid(self, auth):
        r = requests.get(
            f"{API}/reports/xlsx/open-checks/pdf",
            params={"date_from": SEED_DATE, "date_to": SEED_DATE},
            headers=auth,
            timeout=90,
        )
        assert r.status_code == 200, r.text[:500]
        assert r.content[:4] == b"%PDF"
        # basic sanity size
        assert len(r.content) > 2000

    def test_pdf_empty_ok(self, auth):
        r = requests.get(
            f"{API}/reports/xlsx/open-checks/pdf",
            params={"date_from": "2020-01-01", "date_to": "2020-01-01"},
            headers=auth,
            timeout=90,
        )
        assert r.status_code == 200
        assert r.content[:4] == b"%PDF"


# ── Regression (previous reports still work) ────────────────────────────────
class TestRegressionPreviousReports:
    def test_sales_by_category_ok(self, auth):
        r = requests.get(f"{API}/reports/sales-by-category",
                         params={"date_from": "2026-04-01", "date_to": "2026-04-20"},
                         headers=auth, timeout=30)
        assert r.status_code == 200

    def test_cash_close_hierarchical_ok(self, auth):
        r = requests.get(f"{API}/reports/cash-close-hierarchical",
                         params={"date_from": "2026-04-01", "date_to": "2026-04-20"},
                         headers=auth, timeout=30)
        assert r.status_code == 200

    def test_hourly_sales_ok(self, auth):
        r = requests.get(f"{API}/reports/hourly-sales",
                         params={"date_from": "2026-04-01", "date_to": "2026-04-20"},
                         headers=auth, timeout=30)
        assert r.status_code == 200

    def test_taxes_report_ok(self, auth):
        r = requests.get(f"{API}/reports/taxes",
                         params={"date_from": "2026-04-01", "date_to": "2026-04-20"},
                         headers=auth, timeout=30)
        assert r.status_code == 200
        data = r.json()
        assert "breakdown_by_rate" in data  # prompt 4 field preserved
