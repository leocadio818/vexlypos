"""Backend tests for A7 Product Mix by Employee and A4 Sales by Weekday reports."""
import os
import io
import pytest
import requests
from openpyxl import load_workbook

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "https://factura-consumo-app.preview.emergentagent.com").rstrip("/")
ADMIN_PIN = "11338585"
DATE_FROM = "2026-04-01"
DATE_TO = "2026-04-20"


@pytest.fixture(scope="module")
def token():
    r = requests.post(f"{BASE_URL}/api/auth/login", json={"pin": ADMIN_PIN}, timeout=20)
    assert r.status_code == 200, f"login failed: {r.status_code} {r.text[:200]}"
    tk = r.json().get("token")
    assert tk
    return tk


@pytest.fixture(scope="module")
def headers(token):
    return {"Authorization": f"Bearer {token}"}


# ---------- A7 Product Mix by Employee ----------
class TestProductMixByEmployee:
    def test_json_endpoint(self, headers):
        r = requests.get(
            f"{BASE_URL}/api/reports/product-mix-by-employee",
            params={"date_from": DATE_FROM, "date_to": DATE_TO},
            headers=headers,
            timeout=30,
        )
        assert r.status_code == 200, r.text[:300]
        data = r.json()
        assert "employees" in data
        assert "summary" in data
        summary = data["summary"]
        assert abs(float(summary["grand_total"]) - 25100) < 1.0, f"grand_total={summary['grand_total']}"
        assert summary["employee_count"] == 2, f"employee_count={summary.get('employee_count')}"
        emps = data["employees"]
        assert len(emps) == 2
        # Admin should be top (14,400) with PRESIDENTE as top product
        admin = next((e for e in emps if e.get("name") == "Admin"), None)
        carlos = next((e for e in emps if e.get("name") == "Carlos Mesero"), None)
        assert admin, f"Admin not found in {[e.get('name') for e in emps]}"
        assert carlos, f"Carlos Mesero not found"
        admin_total = sum(p["total"] for p in admin["products"])
        carlos_total = sum(p["total"] for p in carlos["products"])
        assert abs(admin_total - 14400) < 1.0, f"admin_total={admin_total}"
        assert abs(carlos_total - 10700) < 1.0, f"carlos_total={carlos_total}"
        # Top product per employee
        assert admin["products"][0]["name"] == "PRESIDENTE"
        assert carlos["products"][0]["name"] == "HAMBURGUESA"

    def test_pdf_download(self, headers):
        r = requests.get(
            f"{BASE_URL}/api/reports/xlsx/product-mix-by-employee/pdf",
            params={"date_from": DATE_FROM, "date_to": DATE_TO},
            headers=headers,
            timeout=60,
        )
        assert r.status_code == 200, r.text[:300]
        assert r.content.startswith(b"%PDF"), f"not PDF header: {r.content[:10]}"

    def test_pdf_requires_auth(self):
        r = requests.get(
            f"{BASE_URL}/api/reports/xlsx/product-mix-by-employee/pdf",
            params={"date_from": DATE_FROM, "date_to": DATE_TO},
            timeout=20,
        )
        assert r.status_code in (401, 403), f"{r.status_code}: {r.text[:200]}"

    def test_xlsx_requires_auth(self):
        r = requests.get(
            f"{BASE_URL}/api/reports/xlsx/product-mix-by-employee/xlsx",
            params={"date_from": DATE_FROM, "date_to": DATE_TO},
            timeout=20,
        )
        assert r.status_code in (401, 403), f"{r.status_code}: {r.text[:200]}"

    def test_xlsx_download_and_formulas(self, headers):
        r = requests.get(
            f"{BASE_URL}/api/reports/xlsx/product-mix-by-employee/xlsx",
            params={"date_from": DATE_FROM, "date_to": DATE_TO},
            headers=headers,
            timeout=60,
        )
        assert r.status_code == 200
        assert r.content[:2] == b"PK"

        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active
        total_employee_rows = []
        grand_total_row = None
        product_rows_with_outline = 0
        for row_idx in range(1, ws.max_row + 1):
            row_vals = [ws.cell(row=row_idx, column=c).value for c in range(1, ws.max_column + 1)]
            first = str(row_vals[0] or "")
            up = first.upper()
            if "TOTAL EMPLEADO" in up:
                total_employee_rows.append((row_idx, row_vals))
            if "TOTAL GENERAL" in up or "GRAN TOTAL" in up:
                grand_total_row = (row_idx, row_vals)
            rd = ws.row_dimensions.get(row_idx)
            if rd and rd.outlineLevel == 1:
                product_rows_with_outline += 1

        assert total_employee_rows, "No TOTAL EMPLEADO rows found"
        assert grand_total_row, "No TOTAL GENERAL row found"
        assert product_rows_with_outline > 0, "No outlineLevel=1 rows (product rows)"

        # Total employee rows must have =SUM(B...) and =SUM(C...)
        for _idx, vals in total_employee_rows:
            sum_b = any(isinstance(v, str) and v.startswith("=SUM(B") for v in vals)
            sum_c = any(isinstance(v, str) and v.startswith("=SUM(C") for v in vals)
            assert sum_b, f"row {_idx} missing =SUM(B..): {vals}"
            assert sum_c, f"row {_idx} missing =SUM(C..): {vals}"

        # Grand total should have SUM formulas referencing discontiguous cells
        gt_vals = grand_total_row[1]
        gt_sums = [v for v in gt_vals if isinstance(v, str) and v.startswith("=SUM(")]
        assert len(gt_sums) >= 2, f"Grand total should have >=2 SUM formulas: {gt_vals}"


# ---------- A4 Sales by Weekday ----------
class TestSalesByWeekday:
    def test_json_endpoint(self, headers):
        r = requests.get(
            f"{BASE_URL}/api/reports/sales-by-weekday",
            params={"date_from": DATE_FROM, "date_to": DATE_TO},
            headers=headers,
            timeout=30,
        )
        assert r.status_code == 200, r.text[:300]
        data = r.json()
        assert "rows" in data
        rows = data["rows"]
        assert len(rows) == 7, f"expected 7 rows, got {len(rows)}"
        # Verify Mon..Sun order
        expected_order = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
        names = [r.get("weekday") or r.get("name") for r in rows]
        assert names == expected_order, f"weekday order wrong: {names}"
        s = data["summary"]
        assert s["peak_weekday"] == "Martes"
        assert abs(float(s["peak_total"]) - 18176) < 1.0
        assert s["valley_weekday"] == "Jueves"
        assert abs(float(s["valley_total"]) - 384) < 1.0
        assert s["best_avg_weekday"] == "Domingo"
        assert abs(float(s["best_avg_per_day"]) - 13584) < 1.0
        assert abs(float(s["grand_total"]) - 32144) < 1.0
        assert int(s["grand_bills"]) == 42

    def test_pdf_download(self, headers):
        r = requests.get(
            f"{BASE_URL}/api/reports/xlsx/sales-by-weekday/pdf",
            params={"date_from": DATE_FROM, "date_to": DATE_TO},
            headers=headers,
            timeout=60,
        )
        assert r.status_code == 200
        assert r.content.startswith(b"%PDF")

    def test_pdf_requires_auth(self):
        r = requests.get(
            f"{BASE_URL}/api/reports/xlsx/sales-by-weekday/pdf",
            params={"date_from": DATE_FROM, "date_to": DATE_TO},
            timeout=20,
        )
        assert r.status_code in (401, 403), r.status_code

    def test_xlsx_requires_auth(self):
        r = requests.get(
            f"{BASE_URL}/api/reports/xlsx/sales-by-weekday/xlsx",
            params={"date_from": DATE_FROM, "date_to": DATE_TO},
            timeout=20,
        )
        assert r.status_code in (401, 403), r.status_code

    def test_xlsx_formulas(self, headers):
        r = requests.get(
            f"{BASE_URL}/api/reports/xlsx/sales-by-weekday/xlsx",
            params={"date_from": DATE_FROM, "date_to": DATE_TO},
            headers=headers,
            timeout=60,
        )
        assert r.status_code == 200
        assert r.content[:2] == b"PK"

        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active
        total_row = None
        ticket_if_count = 0
        avg_per_day_if_count = 0
        pct_formula_count = 0
        for row_idx in range(1, ws.max_row + 1):
            row_vals = [ws.cell(row=row_idx, column=c).value for c in range(1, ws.max_column + 1)]
            first = str(row_vals[0] or "")
            if first.upper() == "TOTAL":
                total_row = (row_idx, row_vals)
            for v in row_vals:
                if not isinstance(v, str):
                    continue
                vx = v.replace(" ", "")
                if vx.startswith("=IF(B") and ">0,C" in vx and "/B" in vx:
                    ticket_if_count += 1
                if vx.startswith("=IF(E") and ">0,C" in vx and "/E" in vx:
                    avg_per_day_if_count += 1
                if vx.startswith("=IF(C") and "/C" in vx:
                    pct_formula_count += 1

        assert total_row, "TOTAL row not found"
        t = total_row[1]
        sums = [v for v in t if isinstance(v, str) and v.startswith("=SUM(")]
        # Expect at least SUM(B..B), SUM(C..C), SUM(E..E)
        assert len(sums) >= 3, f"Expected >=3 SUM formulas in TOTAL row: {t}"
        # Ticket Promedio + Promedio por Día IF formulas on each of 7 rows + total = >=7
        assert ticket_if_count >= 7, f"Expected >=7 ticket IF formulas, got {ticket_if_count}"
        assert avg_per_day_if_count >= 7, f"Expected >=7 avg/day IF formulas, got {avg_per_day_if_count}"
        # % del Total: one per weekday
        assert pct_formula_count >= 7, f"Expected >=7 pct formulas, got {pct_formula_count}"
