"""
Backend tests for Prompt 3: Ventas por Hora report.
Covers:
  - JSON /api/reports/hourly-sales (24 hour buckets)
  - XLSX /api/reports/xlsx/hourly-sales/xlsx (with SUM + IF formulas)
  - PDF  /api/reports/xlsx/hourly-sales/pdf (WeasyPrint)
"""
import os
import io
import re
import pytest
import requests
from openpyxl import load_workbook

BASE_URL = os.environ.get(
    "REACT_APP_BACKEND_URL",
    "https://factura-consumo-app.preview.emergentagent.com",
).rstrip("/")
ADMIN_PIN = "11338585"
DATE_FROM = "2026-04-01"
DATE_TO = "2026-04-20"


@pytest.fixture(scope="module")
def auth_token():
    r = requests.post(f"{BASE_URL}/api/auth/login", json={"pin": ADMIN_PIN}, timeout=15)
    assert r.status_code == 200, f"Login failed: {r.status_code} {r.text}"
    tok = r.json().get("token")
    assert tok
    return tok


@pytest.fixture(scope="module")
def auth_headers(auth_token):
    return {"Authorization": f"Bearer {auth_token}"}


# ---------- JSON endpoint ----------
class TestHourlySalesJSON:
    def test_returns_24_buckets(self, auth_headers):
        r = requests.get(
            f"{BASE_URL}/api/reports/hourly-sales",
            params={"date_from": DATE_FROM, "date_to": DATE_TO},
            headers=auth_headers,
            timeout=20,
        )
        assert r.status_code == 200
        data = r.json()
        assert isinstance(data, list)
        assert len(data) == 24
        for row in data:
            assert set(row.keys()) >= {"hour", "total", "bills"}
            assert re.match(r"^\d{2}:00$", row["hour"])
        peak = max(data, key=lambda x: x["total"])
        assert peak["hour"] == "23:00"
        assert peak["total"] == 7936.0
        assert peak["bills"] == 9


# ---------- XLSX endpoint ----------
class TestHourlySalesXLSX:
    def test_xlsx_formulas(self, auth_headers):
        r = requests.get(
            f"{BASE_URL}/api/reports/xlsx/hourly-sales/xlsx",
            params={"date_from": DATE_FROM, "date_to": DATE_TO},
            headers=auth_headers,
            timeout=30,
        )
        assert r.status_code == 200
        assert r.content[:2] == b"PK"
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active
        # header area + 24 rows + TOTAL GENERAL
        rows = list(ws.iter_rows(values_only=False))
        # Find TOTAL GENERAL row
        total_row = None
        for row in rows:
            if row[0].value == "TOTAL GENERAL":
                total_row = row
                break
        assert total_row is not None, "TOTAL GENERAL row not found"
        # B and C must be SUM formulas
        assert isinstance(total_row[1].value, str) and total_row[1].value.startswith("=SUM(")
        assert isinstance(total_row[2].value, str) and total_row[2].value.startswith("=SUM(")
        # Count hour range rows
        hour_rows = [r for r in rows if isinstance(r[0].value, str) and re.match(r"^\d{2}:00-\d{2}:00$", r[0].value)]
        assert len(hour_rows) == 24
        # Ticket promedio (D) and % (E) are formulas with IF
        sample = hour_rows[0]
        assert isinstance(sample[3].value, str) and sample[3].value.startswith("=IF(")
        assert isinstance(sample[4].value, str) and sample[4].value.startswith("=IF(")
        # % column should reference TOTAL row cell, not hardcoded
        for hr in hour_rows:
            assert "/C" in hr[4].value  # references column C of total row


# ---------- PDF endpoint ----------
class TestHourlySalesPDF:
    def test_pdf_valid_and_content(self, auth_headers):
        r = requests.get(
            f"{BASE_URL}/api/reports/xlsx/hourly-sales/pdf",
            params={"date_from": DATE_FROM, "date_to": DATE_TO},
            headers=auth_headers,
            timeout=60,
        )
        assert r.status_code == 200
        assert r.content[:5] == b"%PDF-"
        try:
            from pypdf import PdfReader
        except ImportError:
            pytest.skip("pypdf not installed")
        reader = PdfReader(io.BytesIO(r.content))
        text = "".join(p.extract_text() for p in reader.pages)
        assert "Ventas por Hora" in text
        assert "TOTAL GENERAL" in text
        assert DATE_FROM in text and DATE_TO in text
        assert "RNC" in text
        ranges = re.findall(r"\d{2}:\d{2}-\d{2}:\d{2}", text)
        assert len(ranges) == 24


# ---------- Auth / error handling ----------
class TestHourlySalesAuth:
    def test_xlsx_requires_auth(self):
        r = requests.get(
            f"{BASE_URL}/api/reports/xlsx/hourly-sales/xlsx",
            params={"date_from": DATE_FROM, "date_to": DATE_TO},
            timeout=15,
        )
        assert r.status_code in (401, 403)

    def test_pdf_requires_auth(self):
        r = requests.get(
            f"{BASE_URL}/api/reports/xlsx/hourly-sales/pdf",
            params={"date_from": DATE_FROM, "date_to": DATE_TO},
            timeout=15,
        )
        assert r.status_code in (401, 403)
