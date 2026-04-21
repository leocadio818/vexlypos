"""Backend tests for A8 — Sales Comparative (Período A vs B).
Covers:
- GET /api/reports/sales-comparative with real dataset
- Auth Bearer enforcement on /xlsx/sales-comparative/{xlsx,pdf}
- XLSX formulas verification (col D =C-B, col E =IF(B=0,"N/A",(C-B)/B))
- Edge case: period_a.total_sales=0 -> pct=null
- Regression: /api/reports/daily-sales still operative
"""
import os
import io
import pytest
import requests
from openpyxl import load_workbook

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL")
if not BASE_URL:
    # Fallback: read frontend/.env (tests invoked outside shell context)
    try:
        with open("/app/frontend/.env") as f:
            for line in f:
                if line.startswith("REACT_APP_BACKEND_URL="):
                    BASE_URL = line.split("=", 1)[1].strip()
                    break
    except Exception:
        pass
BASE_URL = (BASE_URL or "").rstrip("/")
assert BASE_URL, "REACT_APP_BACKEND_URL not configured"
ADMIN_PIN = "11338585"


@pytest.fixture(scope="module")
def admin_token():
    r = requests.post(f"{BASE_URL}/api/auth/login", json={"pin": ADMIN_PIN}, timeout=20)
    assert r.status_code == 200, f"Login failed: {r.status_code} {r.text[:200]}"
    return r.json().get("token")


@pytest.fixture(scope="module")
def auth_headers(admin_token):
    return {"Authorization": f"Bearer {admin_token}"}


# ─── Core endpoint: /api/reports/sales-comparative ───
class TestSalesComparativeEndpoint:
    def test_real_dataset_deltas(self, auth_headers):
        params = {
            "period_a_from": "2026-04-01", "period_a_to": "2026-04-10",
            "period_b_from": "2026-04-11", "period_b_to": "2026-04-20",
        }
        r = requests.get(f"{BASE_URL}/api/reports/sales-comparative",
                         params=params, headers=auth_headers, timeout=30)
        assert r.status_code == 200, r.text[:300]
        data = r.json()

        # Schema
        assert "period_a" in data and "period_b" in data and "metrics" in data
        assert isinstance(data["metrics"], list) and len(data["metrics"]) == 9

        pa, pb = data["period_a"], data["period_b"]
        # Expected values per review request
        assert pa["total_sales"] == 18304, f"period_a.total_sales={pa['total_sales']}"
        assert pa["bills_count"] == 19, f"period_a.bills_count={pa['bills_count']}"
        assert pb["total_sales"] == 13840, f"period_b.total_sales={pb['total_sales']}"
        assert pb["bills_count"] == 23, f"period_b.bills_count={pb['bills_count']}"

        # Metric deltas
        by_key = {m["key"]: m for m in data["metrics"]}
        assert by_key["total_sales"]["diff"] == -4464
        assert by_key["total_sales"]["pct"] == -24.39
        assert by_key["bills_count"]["diff"] == 4
        assert by_key["bills_count"]["pct"] == 21.05

        # Metric keys present (9)
        expected_keys = {"total_sales", "bills_count", "avg_ticket", "subtotal",
                         "total_itbis", "total_tips", "total_discount", "cash_sales", "card_sales"}
        assert set(by_key.keys()) == expected_keys

    def test_pct_null_when_period_a_is_zero(self, auth_headers):
        """Use an empty A range (year 2019) so totals are 0 -> pct must be None."""
        params = {
            "period_a_from": "2019-01-01", "period_a_to": "2019-01-02",
            "period_b_from": "2026-04-11", "period_b_to": "2026-04-20",
        }
        r = requests.get(f"{BASE_URL}/api/reports/sales-comparative",
                         params=params, headers=auth_headers, timeout=30)
        assert r.status_code == 200
        data = r.json()
        assert data["period_a"]["total_sales"] == 0
        ts_metric = next(m for m in data["metrics"] if m["key"] == "total_sales")
        assert ts_metric["pct"] is None, f"pct should be null when period_a=0; got {ts_metric['pct']}"


# ─── XLSX/PDF downloads ───
class TestSalesComparativeExports:
    params = {
        "period_a_from": "2026-04-01", "period_a_to": "2026-04-10",
        "period_b_from": "2026-04-11", "period_b_to": "2026-04-20",
    }

    def test_xlsx_requires_auth(self):
        r = requests.get(f"{BASE_URL}/api/reports/xlsx/sales-comparative/xlsx",
                         params=self.params, timeout=30)
        assert r.status_code in (401, 403), f"Expected auth required, got {r.status_code}"

    def test_pdf_requires_auth(self):
        r = requests.get(f"{BASE_URL}/api/reports/xlsx/sales-comparative/pdf",
                         params=self.params, timeout=30)
        assert r.status_code in (401, 403), f"Expected auth required, got {r.status_code}"

    def test_xlsx_download_valid(self, auth_headers):
        r = requests.get(f"{BASE_URL}/api/reports/xlsx/sales-comparative/xlsx",
                         params=self.params, headers=auth_headers, timeout=60)
        assert r.status_code == 200, r.text[:300]
        assert r.content[:2] == b"PK", "XLSX must start with PK"
        assert len(r.content) > 1000

    def test_pdf_download_valid(self, auth_headers):
        r = requests.get(f"{BASE_URL}/api/reports/xlsx/sales-comparative/pdf",
                         params=self.params, headers=auth_headers, timeout=90)
        assert r.status_code == 200, r.text[:300]
        assert r.content[:4] == b"%PDF", "PDF must start with %PDF"
        assert len(r.content) > 5000

    def test_xlsx_contains_real_formulas(self, auth_headers):
        """Validate column D (Diferencia) = '=C{row}-B{row}' (formula, not value)
           and column E (% Cambio) = '=IF(B{row}=0,"N/A",(C{row}-B{row})/B{row})'."""
        r = requests.get(f"{BASE_URL}/api/reports/xlsx/sales-comparative/xlsx",
                         params=self.params, headers=auth_headers, timeout=60)
        assert r.status_code == 200
        wb = load_workbook(io.BytesIO(r.content))  # default: data_only=False, keeps formulas
        ws = wb.active
        # Metrics rows start at row 7 (header row 6 per impl)
        checked = 0
        for row in range(7, 7 + 9):
            d_val = ws.cell(row=row, column=4).value
            e_val = ws.cell(row=row, column=5).value
            assert isinstance(d_val, str) and d_val == f"=C{row}-B{row}", \
                f"Row {row} col D expected formula =C{row}-B{row}, got {d_val!r}"
            assert isinstance(e_val, str) and e_val == f'=IF(B{row}=0,"N/A",(C{row}-B{row})/B{row})', \
                f"Row {row} col E expected formula, got {e_val!r}"
            checked += 1
        assert checked == 9


# ─── Regression: legacy daily-sales still works ───
class TestDailySalesRegression:
    def test_daily_sales_still_works(self, auth_headers):
        r = requests.get(f"{BASE_URL}/api/reports/daily-sales",
                         params={"date_from": "2026-04-01", "date_to": "2026-04-10"},
                         headers=auth_headers, timeout=30)
        assert r.status_code == 200, r.text[:200]
        data = r.json()
        # Should have the shape that _sales_snapshot consumes
        for k in ("total_sales", "bills_count", "avg_ticket"):
            assert k in data, f"daily-sales missing '{k}'"

    def test_open_checks_regression(self, auth_headers):
        """B5 Open Checks regression check."""
        r = requests.get(f"{BASE_URL}/api/reports/open-checks",
                         headers=auth_headers, timeout=30)
        # Acceptable 200 (impl) or 422 if params are required — we treat 5xx/404 as failure
        assert r.status_code in (200, 422), f"open-checks returned {r.status_code}"
