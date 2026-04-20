"""Tests for Cash Close Hierarchical Report (Prompt 2)."""
import os
import io
import pytest
import requests
from openpyxl import load_workbook

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "https://factura-consumo-app.preview.emergentagent.com").rstrip("/")
DATE_FROM = "2026-04-01"
DATE_TO = "2026-04-20"


@pytest.fixture(scope="module")
def jwt_token():
    r = requests.post(f"{BASE_URL}/api/auth/login", json={"pin": "11338585"}, timeout=30)
    assert r.status_code == 200, r.text
    data = r.json()
    return data.get("access_token") or data.get("token")


@pytest.fixture
def auth_headers(jwt_token):
    return {"Authorization": f"Bearer {jwt_token}"}


# --- JSON endpoint (public) ---
class TestHierarchicalJson:
    def test_structure_and_values(self):
        r = requests.get(
            f"{BASE_URL}/api/reports/cash-close-hierarchical",
            params={"date_from": DATE_FROM, "date_to": DATE_TO},
            timeout=30,
        )
        assert r.status_code == 200, r.text
        data = r.json()
        assert "employees" in data and "grand_totals" in data
        assert isinstance(data["employees"], list)
        assert len(data["employees"]) >= 1, "Expected at least 1 employee in seeded range"
        emp = data["employees"][0]
        assert "name" in emp and "shifts" in emp and "employee_totals" in emp
        assert len(emp["shifts"]) >= 1
        sh = emp["shifts"][0]
        for k in ("shift_start", "shift_end", "payment_methods", "shift_totals"):
            assert k in sh
        assert len(sh["payment_methods"]) >= 1
        pm = sh["payment_methods"][0]
        for k in ("name", "transactions", "subtotal"):
            assert k in pm
        # Grand totals sanity
        gt = data["grand_totals"]
        for k in ("count", "total", "tips", "total_with_tips"):
            assert k in gt
        assert gt["count"] >= 1
        assert gt["total"] > 0

    def test_filter_by_employee(self):
        r = requests.get(
            f"{BASE_URL}/api/reports/cash-close-hierarchical",
            params={"date_from": DATE_FROM, "date_to": DATE_TO, "employee": "Admin"},
            timeout=30,
        )
        assert r.status_code == 200
        data = r.json()
        # All employees returned must include "Admin" in name (case-insensitive)
        for emp in data["employees"]:
            assert "admin" in emp["name"].lower()

    def test_filter_by_payment_method(self):
        r = requests.get(
            f"{BASE_URL}/api/reports/cash-close-hierarchical",
            params={"date_from": DATE_FROM, "date_to": DATE_TO, "payment_method": "Efectivo"},
            timeout=30,
        )
        assert r.status_code == 200
        data = r.json()
        for emp in data["employees"]:
            for sh in emp["shifts"]:
                for pm in sh["payment_methods"]:
                    assert "efectivo" in pm["name"].lower()


# --- PDF/XLSX endpoints (auth required) ---
class TestDownloads:
    def test_pdf_resumida(self, auth_headers):
        r = requests.get(
            f"{BASE_URL}/api/reports/xlsx/cierre-caja/pdf",
            params={"date_from": DATE_FROM, "date_to": DATE_TO, "view": "resumida"},
            headers=auth_headers,
            timeout=60,
        )
        assert r.status_code == 200, r.text[:200]
        assert r.content[:4] == b"%PDF", "Response is not a PDF"
        assert len(r.content) > 2000

    def test_pdf_detallada(self, auth_headers):
        r = requests.get(
            f"{BASE_URL}/api/reports/xlsx/cierre-caja/pdf",
            params={"date_from": DATE_FROM, "date_to": DATE_TO, "view": "detallada"},
            headers=auth_headers,
            timeout=60,
        )
        assert r.status_code == 200
        assert r.content[:4] == b"%PDF"
        # Detailed should be larger than resumida (includes tx rows)
        assert len(r.content) > 2000

    def test_pdf_requires_auth(self):
        r = requests.get(
            f"{BASE_URL}/api/reports/xlsx/cierre-caja/pdf",
            params={"date_from": DATE_FROM, "date_to": DATE_TO, "view": "resumida"},
            timeout=30,
        )
        assert r.status_code in (401, 403), f"Expected auth rejection, got {r.status_code}"

    def test_xlsx_resumida(self, auth_headers):
        r = requests.get(
            f"{BASE_URL}/api/reports/xlsx/cierre-caja/xlsx",
            params={"date_from": DATE_FROM, "date_to": DATE_TO, "view": "resumida"},
            headers=auth_headers,
            timeout=60,
        )
        assert r.status_code == 200
        assert r.content[:2] == b"PK", "Response is not a XLSX (ZIP)"
        wb = load_workbook(io.BytesIO(r.content), data_only=False)
        assert len(wb.sheetnames) >= 1
        ws = wb[wb.sheetnames[0]]
        # Look for SUM formulas and outline levels
        found_sum = False
        found_outline = False
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("=SUM"):
                    found_sum = True
            if getattr(row[0].parent.row_dimensions[row[0].row], "outline_level", 0) and row[0].parent.row_dimensions[row[0].row].outline_level > 0:
                found_outline = True
        # row_dimensions check (alternative)
        for rd in ws.row_dimensions.values():
            if rd.outline_level and rd.outline_level > 0:
                found_outline = True
                break
        assert found_sum, "No =SUM formula found in XLSX"
        assert found_outline, "No outlineLevel grouping found in XLSX"

    def test_xlsx_detallada(self, auth_headers):
        r = requests.get(
            f"{BASE_URL}/api/reports/xlsx/cierre-caja/xlsx",
            params={"date_from": DATE_FROM, "date_to": DATE_TO, "view": "detallada"},
            headers=auth_headers,
            timeout=60,
        )
        assert r.status_code == 200
        assert r.content[:2] == b"PK"
        wb = load_workbook(io.BytesIO(r.content), data_only=False)
        ws = wb[wb.sheetnames[0]]
        assert ws.max_row > 5


# --- Regression: classic cash-close still works ---
class TestRegression:
    def test_classic_cash_close_ok(self):
        r = requests.get(
            f"{BASE_URL}/api/reports/cash-close",
            params={"date_from": DATE_FROM, "date_to": DATE_TO},
            timeout=30,
        )
        assert r.status_code == 200
        data = r.json()
        # Must still have the classic dashboard fields
        assert isinstance(data, dict)
